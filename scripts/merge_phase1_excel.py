"""
Phase 1: Extract and merge Excel-based 资金用途调整表 files.
Outputs:
  - data/celma-major-events-attachments/merged_用途调整表.xlsx  (merged standard-format rows)
  - data/celma-major-events-attachments/temp/classify_results.json  (updated with per-file status)
"""
import json, os, re, sys
import xlrd
import openpyxl
from openpyxl import Workbook

TEMP = 'data/celma-major-events-attachments/temp'
CLASSIFY = os.path.join(TEMP, 'classify_results.json')
OUT_XLSX = 'data/celma-major-events-attachments/merged_用途调整表.xlsx'

# Standard 30-column header (the canonical schema after flattening merged cells)
STD_COLS = [
    '序号',
    '债券编码', '债券简称', '债券全称', '发行日期', '到期日期',
    '发行利率', '发行金额', '未到期金额', '未使用金额', '用途调整金额',
    '市县名称（调整前）', '区划编码（调整前）', '市县名称（调整后）', '区划编码（调整后）',
    '调整前项目名称', '调整前项目编码', '调整前项目领域', '调整前主管部门',
    '调整前项目单位', '调整前建设状态', '调整原因',
    '调整后项目名称', '调整后项目编码', '调整后项目领域', '调整后主管部门',
    '调整后项目单位', '调整后建设状态', '建设期限', '预计竣工日期',
    '备注',
]

# Keywords to identify the header row
HEADER_KEYWORDS = {'债券编码', '债券简称', '发行日期', '用途调整', '项目名称', '项目编码'}
# Keywords to skip summary/total rows
SKIP_KEYWORDS = {'合计', '全区合计', '全省合计', '全市合计', '总计', '小计'}


def clean_cell(v):
    """Clean a cell value: strip whitespace/newlines, convert to string."""
    if v is None:
        return ''
    s = str(v).strip().replace('\n', ' ').replace('\r', '')
    # Remove trailing .0 from numbers that were read as floats
    if re.match(r'^\d+\.0$', s):
        s = s[:-2]
    return s


def find_header_row(rows, max_scan=10):
    """Find the row index that contains the actual column headers (debt code, etc.)."""
    for i, row in enumerate(rows[:max_scan]):
        text = ' '.join(clean_cell(c) for c in row).lower()
        # The header row should contain "债券编码" or similar
        if '债券编码' in text or '债券 编码' in text:
            return i
    return -1


def is_skip_row(row):
    """Check if this is a summary/total row that should be skipped."""
    first_cells = ' '.join(clean_cell(c) for c in row[:3])
    for kw in SKIP_KEYWORDS:
        if kw in first_cells:
            return True
    # Skip empty rows
    non_empty = sum(1 for c in row if clean_cell(c))
    if non_empty <= 1:
        return True
    return False


def detect_format(header_row, ncols):
    """Detect if this is standard format A (~30 cols) or simplified format B."""
    text = ' '.join(clean_cell(c) for c in header_row)
    has_bond_code = '债券编码' in text
    has_project_code = '项目编码' in text
    has_adjust_amount = '调整金额' in text or '用途调整' in text

    if ncols >= 28 and has_bond_code and has_project_code:
        return 'standard'
    elif ncols >= 20 and has_bond_code:
        return 'standard'  # Slightly narrower but still standard-ish
    else:
        return 'simplified'


def normalize_standard_row(row, ncols):
    """Normalize a standard-format row to exactly 31 columns (STD_COLS)."""
    cells = [clean_cell(c) for c in row]
    # Pad or trim to 31
    if len(cells) < 31:
        cells.extend([''] * (31 - len(cells)))
    elif len(cells) > 31:
        # Some files have 32 cols (extra "需求申报年度" before 调整前/后项目)
        # Or extra trailing column. Just trim extra.
        cells = cells[:31]
    return cells


def read_xls(path):
    """Read .xls file, return list of rows (each row is list of values)."""
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)
    rows = []
    for r in range(sh.nrows):
        rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
    return rows, sh.ncols


def read_xlsx(path):
    """Read .xlsx file, return list of rows."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb.active
    rows = []
    max_col = sh.max_column or 0
    for row in sh.iter_rows(values_only=True):
        rows.append(list(row))
    return rows, max_col


def has_extra_year_cols(header_row):
    """Check if the header has '需求申报年度' columns (山西 format - 32 cols)."""
    text = ' '.join(clean_cell(c) for c in header_row)
    return '需求申报年度' in text


def remap_shanxi_row(row):
    """Remap 山西-style 32-col rows (with 需求申报年度) to standard 31-col."""
    cells = [clean_cell(c) for c in row]
    if len(cells) < 32:
        cells.extend([''] * (32 - len(cells)))
    # 山西 layout: cols 0-14 same, then col15=需求申报年度, cols16-22=调整前, col23=需求申报年度, cols24-31=调整后
    # Standard: cols 0-14 same, cols15-21=调整前(no 需求申报年度), col22=调整原因, cols23-30=调整后
    # Remove the two 需求申报年度 columns (15 and 23-after-removal)
    out = cells[:15] + cells[16:23] + cells[24:32]
    if len(out) < 31:
        out.extend([''] * (31 - len(out)))
    return out[:31]


def process_excel_file(entry, classify_entry):
    """Process a single Excel file. Returns (data_rows, format_type, issues)."""
    path = entry['path']
    folder = entry['folder']
    name = entry['name']
    folder_name = os.path.basename(folder)

    try:
        if path.endswith('.xls'):
            rows, ncols = read_xls(path)
        else:
            rows, ncols = read_xlsx(path)
    except Exception as e:
        classify_entry.update({
            'mergeable': False,
            'format': 'error',
            'issue': f'读取失败: {e}',
        })
        return [], 'error', str(e)

    if len(rows) < 3:
        classify_entry.update({
            'mergeable': False,
            'format': 'too_few_rows',
            'issue': f'行数过少: {len(rows)}',
        })
        return [], 'too_few_rows', 'too few rows'

    hdr_idx = find_header_row(rows)
    if hdr_idx < 0:
        classify_entry.update({
            'mergeable': False,
            'format': 'no_header',
            'issue': '未找到表头行（无"债券编码"等关键词）',
        })
        return [], 'no_header', 'header not found'

    header_row = rows[hdr_idx]
    fmt = detect_format(header_row, ncols)

    if fmt == 'simplified':
        # Save raw header for reference
        raw_hdr = [clean_cell(c) for c in header_row]
        classify_entry.update({
            'mergeable': False,
            'format': 'simplified',
            'issue': f'简化格式 ({ncols}列): {", ".join(c for c in raw_hdr if c)[:120]}',
            'ncols': ncols,
        })
        return [], 'simplified', 'simplified format'

    # Standard format - check for 山西 variant with 需求申报年度
    is_shanxi = has_extra_year_cols(header_row)

    data_rows = []
    issues = []
    for i in range(hdr_idx + 1, len(rows)):
        row = rows[i]
        if is_skip_row(row):
            continue

        if is_shanxi:
            cells = remap_shanxi_row(row)
        else:
            cells = normalize_standard_row(row, ncols)

        # Basic validation: at least some key fields should be non-empty
        bond_code = cells[1]  # 债券编码
        project_name_before = cells[15]  # 调整前项目名称
        project_name_after = cells[22]  # 调整后项目名称
        adjust_amount = cells[10]  # 用途调整金额

        if not bond_code and not project_name_before and not project_name_after and not adjust_amount:
            continue  # Skip truly empty rows

        data_rows.append(cells)

    if len(data_rows) == 0:
        classify_entry.update({
            'mergeable': False,
            'format': 'standard_empty',
            'issue': '标准格式但无有效数据行',
        })
        return [], 'standard_empty', 'no data rows'

    classify_entry.update({
        'mergeable': True,
        'format': 'standard' + ('_shanxi' if is_shanxi else ''),
        'issue': '',
        'data_rows': len(data_rows),
        'ncols': ncols,
    })
    return data_rows, 'standard', ''


def main():
    # Load classify results
    with open(CLASSIFY, 'r', encoding='utf8') as f:
        classify = json.load(f)

    excel_entries = classify.get('excel', [])
    print(f'Total Excel files: {len(excel_entries)}')

    # Initialize per-file tracking
    file_results = []
    all_data_rows = []
    standard_count = 0
    simplified_count = 0
    error_count = 0

    for entry in excel_entries:
        folder_name = os.path.basename(entry['folder'])
        file_name = entry['name']
        classify_entry = {
            'path': entry['path'],
            'folder_name': folder_name,
            'file_name': file_name,
            'file_type': 'excel',
            'source': 'phase1_excel',
        }

        data_rows, fmt, issue = process_excel_file(entry, classify_entry)

        if data_rows:
            # Prepend folder_name and file_name to each row
            for row in data_rows:
                all_data_rows.append([folder_name, file_name] + row)
            standard_count += 1
            print(f'  ✓ {file_name} | {len(data_rows)} rows | {folder_name[:40]}')
        else:
            if fmt == 'simplified':
                simplified_count += 1
                print(f'  ◇ SIMPLIFIED: {file_name} | {folder_name[:40]}')
            else:
                error_count += 1
                print(f'  ✗ {fmt}: {file_name} | {issue[:60]}')

        file_results.append(classify_entry)

    # Write merged output
    print(f'\n=== SUMMARY ===')
    print(f'Standard mergeable: {standard_count} files, {len(all_data_rows)} total rows')
    print(f'Simplified (deferred): {simplified_count} files')
    print(f'Errors: {error_count} files')

    if all_data_rows:
        wb = Workbook()
        ws = wb.active
        ws.title = '资金用途调整表'
        # Write header
        ws.append(['文件夹名称', '表格文件名称'] + STD_COLS)
        for row in all_data_rows:
            ws.append(row)
        wb.save(OUT_XLSX)
        print(f'\nMerged output saved: {OUT_XLSX} ({len(all_data_rows)} data rows)')

    # Update classify_results.json
    classify['excel_results'] = file_results
    classify['excel_summary'] = {
        'total': len(excel_entries),
        'standard_merged': standard_count,
        'simplified_deferred': simplified_count,
        'errors': error_count,
        'total_data_rows': len(all_data_rows),
    }
    with open(CLASSIFY, 'w', encoding='utf8') as f:
        json.dump(classify, f, ensure_ascii=False, indent=2)
    print(f'Updated {CLASSIFY}')


if __name__ == '__main__':
    main()
