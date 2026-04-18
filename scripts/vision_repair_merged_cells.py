"""
Vision repair for merged-cell issues in extracted tables.

Identifies files with poor/medium quality (low mapped_cols, high empty_rate)
and re-extracts them using the vision model with an enhanced prompt that
explicitly instructs the model to expand merged cells.

For textual PDFs: converts PDF pages to images, sends to vision
For scanned PDFs: re-sends images to vision with enhanced merged-cell prompt
For Excel files: reads xlsx with openpyxl (merged_cells aware) to expand merges

Outputs to: temp/extracted_vision_repair/
Then re-merge can pick these up by preferring them over originals.

Run:
  python scripts/vision_repair_merged_cells.py                # process all targets
  python scripts/vision_repair_merged_cells.py --list         # just list targets
  python scripts/vision_repair_merged_cells.py --limit 5      # cap number of files
  python scripts/vision_repair_merged_cells.py --phase textual  # only textual PDFs
  python scripts/vision_repair_merged_cells.py --phase scanned  # only scanned PDFs
  python scripts/vision_repair_merged_cells.py --phase excel    # only Excel files
  python scripts/vision_repair_merged_cells.py --grade very_low # only very_low grade
  python scripts/vision_repair_merged_cells.py --force          # re-process even if done
"""

import json, os, re, sys, base64, io, time, traceback
from collections import Counter

os.chdir(r"D:\Research\202603-自动化债务报告")

TEMP = "data/celma-major-events-attachments/temp"
CLASSIFY = os.path.join(TEMP, "classify_results.json")
REPAIR_DIR = os.path.join(TEMP, "extracted_vision_repair")
TEXT_DIR = os.path.join(TEMP, "extracted_textual")
SCAN_DIR = os.path.join(TEMP, "extracted_scanned")
VISION_DIR = os.path.join(TEMP, "extracted_scanned_vision")
LOG_FILE = os.path.join(TEMP, "vision_repair_log.txt")

DPI = 300
MAX_IMG_DIM = 2400

os.makedirs(REPAIR_DIR, exist_ok=True)

# ──────────── Enhanced merged-cell-aware prompt ────────────

MERGED_CELL_PROMPT = """请从这张图片中提取表格数据。这是一张"地方政府债券资金用途调整表"。

【重要】这类表格普遍存在大量合并单元格：
- 同一债券信息（债券编码、简称、全称、发行日期、到期日期、发行利率、发行金额、未到期金额、未使用金额、用途调整金额）会被合并成一个大单元格，覆盖多行数据
- 同一区划信息（市县名称、区划编码）也可能合并多行
- 你必须将合并单元格的值复制到它覆盖的每一行中，确保每一行数据都是完整的

请严格按照以下规则输出：
1. 每行数据用 | 分隔各列
2. 这是标准31列表格，列顺序为：序号|债券编码|债券简称|债券全称|发行日期|到期日期|发行利率|发行金额|未到期金额|未使用金额|用途调整金额|市县名称（调整前）|区划编码（调整前）|市县名称（调整后）|区划编码（调整后）|调整前项目名称|调整前项目编码|调整前项目领域|调整前主管部门|调整前项目单位|调整前建设状态|调整原因|调整后项目名称|调整后项目编码|调整后项目领域|调整后主管部门|调整后项目单位|调整后建设状态|建设期限|预计竣工日期|备注
3. 对于合并单元格，将合并单元格的内容填充到每一行。例如，如果"债券编码"单元格合并了3行，则这3行都应包含相同的债券编码值
4. 保留原始表头行（但只需保留字段名行，如"序号|债券编码|债券简称..."）
5. 保留所有数字精度，不要四舍五入
6. 空单元格用空字符串表示（两个|之间无内容）
7. 不要添加任何解释文字，只输出表格数据
8. 不要输出"合计"、"小计"等汇总行
9. 如果表头有多行分类行（如"一、债券信息"），可以跳过这些分类行，只保留字段名行

示例输出（注意每行都包含完整信息）：
序号|债券编码|债券简称|债券全称|发行日期|到期日期|发行利率|发行金额|未到期金额|未使用金额|用途调整金额|市县名称|区划编码|市县名称|区划编码|项目名称|项目编码|项目领域|主管部门|项目单位|建设状态|调整原因|项目名称|项目编码|项目领域|主管部门|项目单位|建设状态|建设期限|预计竣工日期|备注
1|2400001|XX专项01|XX省2024年专项债券(一期)|2024-01-15|2034-01-15|2.88|500000|500000|100000|50000|XX市|410100|XX市|410100|旧项目A|P001|交通|交通局|XX公司|在建|项目调整|新项目B|P002|教育|教育局|YY公司|在建|2024-2026|2026-12-31|
2|2400001|XX专项01|XX省2024年专项债券(一期)|2024-01-15|2034-01-15|2.88|500000|500000|100000|30000|YY县|410200|YY县|410200|旧项目C|P003|医疗|卫健委|ZZ单位|完工|资金调整|新项目D|P004|水利|水利局|WW单位|在建|2024-2025|2025-06-30|

注意上面示例中第1行和第2行的债券编码、简称、全称等信息完全相同——这就是合并单元格展开后的正确结果。"""

# ──────────── helpers ────────────

def clean(v):
    if v is None:
        return ""
    return str(v).strip().replace("\n", " ").replace("\r", "")


def safe_sheet_name(name, existing):
    name = re.sub(r'[\[\]:*?/\\]', '_', name)[:28]
    if name in existing:
        for i in range(2, 100):
            n2 = f'{name}_{i}'
            if n2 not in existing:
                return n2
    return name


def prepare_image_bytes(image_bytes):
    """Resize and compress image to stay under API limit."""
    from PIL import Image as PILImage
    img = PILImage.open(io.BytesIO(image_bytes))
    if max(img.size) > MAX_IMG_DIM:
        ratio = MAX_IMG_DIM / max(img.size)
        img = img.resize((int(img.width * ratio), int(img.height * ratio)))
    # Try PNG first
    buf = io.BytesIO()
    img.save(buf, 'PNG')
    out = buf.getvalue()
    media_type = 'image/png'
    if len(out) > 4_500_000:
        for q in [85, 60, 40, 25]:
            buf = io.BytesIO()
            img.convert('RGB').save(buf, 'JPEG', quality=q)
            out = buf.getvalue()
            media_type = 'image/jpeg'
            if len(out) <= 4_500_000:
                break
    return out, media_type


def vision_extract_page(image_bytes, page_num=1, max_tokens=16000):
    """Call vision model for one page with merged-cell-aware prompt.
    Returns (rows: list[list[str]], error: str).
    """
    import anthropic
    import httpx

    api_key = os.environ.get('ANTHROPIC_AUTH_TOKEN') or os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        return [], 'no API key'

    base_url = os.environ.get('ANTHROPIC_BASE_URL', '').rstrip('/')
    kwargs = {'api_key': api_key}
    if base_url:
        kwargs['base_url'] = base_url
    # Set a 5-minute timeout for vision calls (large images take time)
    kwargs['timeout'] = httpx.Timeout(300.0, connect=30.0)
    client = anthropic.Anthropic(**kwargs)

    img_bytes, media_type = prepare_image_bytes(image_bytes)
    b64 = base64.b64encode(img_bytes).decode('utf-8')

    model = os.environ.get('ANTHROPIC_MODEL', 'claude-sonnet-4-6')
    try:
        msg = client.messages.create(
            model=model,
            max_tokens=max_tokens,
            messages=[{
                'role': 'user',
                'content': [
                    {'type': 'image', 'source': {'type': 'base64', 'media_type': media_type, 'data': b64}},
                    {'type': 'text', 'text': MERGED_CELL_PROMPT}
                ]
            }]
        )
        text = msg.content[0].text
        rows = []
        for line in text.strip().split('\n'):
            line = line.strip()
            if line and '|' in line:
                cells = [c.strip() for c in line.split('|')]
                rows.append(cells)
        return rows, ''
    except Exception as e:
        err_str = str(e)[:300]
        return [], err_str


def save_repair_xlsx(pages_data, out_path):
    """Save vision repair results for one file.
    pages_data: list of (page_num, rows, quality_str)
    """
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    sheet_names = set()
    for page_num, rows, quality in pages_data:
        sn = safe_sheet_name(f'p{page_num}_repair_{quality}', sheet_names)
        sheet_names.add(sn)
        ws = wb.create_sheet(title=sn)
        ws.append([f'# page={page_num}, method=vision_repair, quality={quality}'])
        for row in rows:
            ws.append(row)
    if not pages_data:
        ws = wb.create_sheet(title='empty')
        ws.append(['No tables found'])
    wb.save(out_path)


# ──────────── Excel merged cell repair ────────────

def repair_excel_merged_cells(excel_path, out_path):
    """Read Excel file with merged cells awareness and expand them.
    Returns (n_rows, issue_str).
    """
    from openpyxl import load_workbook, Workbook

    try:
        # Open with data_only but NOT read_only (need merged_cells info)
        wb = load_workbook(excel_path, data_only=True, read_only=False)
        ws = wb.active

        # Build merged cell map: (row, col) -> value
        merge_map = {}
        for merge_range in ws.merged_cells.ranges:
            min_row, min_col = merge_range.min_row, merge_range.min_col
            val = ws.cell(row=min_row, column=min_col).value
            for r in range(merge_range.min_row, merge_range.max_row + 1):
                for c in range(merge_range.min_col, merge_range.max_col + 1):
                    if (r, c) != (min_row, min_col):
                        merge_map[(r, c)] = val

        # Read all rows with merged cells expanded
        rows = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            row_data = []
            for c_idx, cell in enumerate(row, start=1):
                val = cell.value
                if val is None and (r_idx, c_idx) in merge_map:
                    val = merge_map[(r_idx, c_idx)]
                row_data.append(clean(val) if val is not None else "")
            rows.append(row_data)
        wb.close()

        if not rows:
            return 0, "empty"

        # Save to output xlsx
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "repaired"
        ws_out.append(["# method=excel_merge_repair"])
        for row in rows:
            ws_out.append(row)
        wb_out.save(out_path)
        return len(rows), ""

    except Exception as e:
        return 0, str(e)[:300]


def repair_xls_merged_cells(xls_path, out_path):
    """Read .xls file with xlrd, detect merged cells, expand them.
    Returns (n_rows, issue_str).
    """
    import xlrd
    from openpyxl import Workbook

    try:
        wb = xlrd.open_workbook(xls_path, formatting_info=True)
        ws = wb.sheet_by_index(0)

        # Build merged cell map from xlrd
        merge_map = {}
        for rlo, rhi, clo, chi in ws.merged_cells:
            val = ws.cell_value(rlo, clo)
            for r in range(rlo, rhi):
                for c in range(clo, chi):
                    if (r, c) != (rlo, clo):
                        merge_map[(r, c)] = val

        rows = []
        for r in range(ws.nrows):
            row_data = []
            for c in range(ws.ncols):
                val = ws.cell_value(r, c)
                if (not val or val == '') and (r, c) in merge_map:
                    val = merge_map[(r, c)]
                row_data.append(clean(val) if val else "")
            rows.append(row_data)
        wb.release_resources()

        if not rows:
            return 0, "empty"

        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "repaired"
        ws_out.append(["# method=excel_merge_repair"])
        for row in rows:
            ws_out.append(row)
        wb_out.save(out_path)
        return len(rows), ""

    except Exception as e:
        return 0, str(e)[:300]


# ──────────── Collect repair targets ────────────

def collect_targets(classify, phase_filter=None, grade_filter=None):
    """Collect repair targets from merge_from_extracted_details."""
    details = classify.get("merge_from_extracted_details", {})
    targets = []

    # Path lookup for scanned PDFs
    path_lookup_scanned = {}
    for entry in classify.get("scanned", []):
        k = entry["name"] + "|" + os.path.basename(entry["folder"])
        path_lookup_scanned[k] = entry["path"]

    # Path lookup for textual PDFs
    path_lookup_textual = {}
    for entry in classify.get("textual", []):
        k = entry["name"] + "|" + os.path.basename(entry["folder"])
        path_lookup_textual[k] = entry["path"]

    # Path lookup for Excel files
    path_lookup_excel = {}
    for entry in classify.get("excel", []):
        k = entry["name"] + "|" + os.path.basename(entry["folder"])
        path_lookup_excel[k] = entry["path"]

    for phase_name in ["textual", "scanned", "excel"]:
        if phase_filter and phase_name != phase_filter:
            continue
        phase_data = details.get(phase_name, {})
        # Handle both old and new format
        if isinstance(phase_data, dict):
            file_list = phase_data.get("files", [])
        elif isinstance(phase_data, list):
            file_list = phase_data
        else:
            continue

        for item in file_list:
            if isinstance(item, str):
                continue
            q = item.get("quality", {})
            if isinstance(q, str):
                continue
            grade = q.get("quality_grade", "")
            if grade not in ("medium", "low", "very_low"):
                continue
            if grade_filter and grade != grade_filter:
                continue

            file_name = item.get("file", "")
            folder_name = item.get("folder", "")
            key = file_name + "|" + folder_name

            # Find source PDF/Excel path
            if phase_name == "textual":
                src_path = path_lookup_textual.get(key, "")
            elif phase_name == "scanned":
                src_path = path_lookup_scanned.get(key, "")
            else:
                src_path = path_lookup_excel.get(key, "")

            # Build output name
            safe_name = re.sub(r'[<>:"|?*]', '_',
                               f'{folder_name}__{os.path.splitext(file_name)[0]}.xlsx')
            out_path = os.path.join(REPAIR_DIR, safe_name)

            targets.append({
                "phase": phase_name,
                "file": file_name,
                "folder": folder_name,
                "key": key,
                "rows": item.get("rows", 0),
                "grade": grade,
                "mapped_cols": q.get("mapped_cols", 0),
                "empty_rate": q.get("empty_rate", 0),
                "src_path": src_path,
                "out_name": safe_name,
                "out_path": out_path,
            })

    return targets


# ──────────── Main ────────────

def main():
    force = '--force' in sys.argv
    list_only = '--list' in sys.argv
    phase_filter = None
    grade_filter = None
    limit = None

    for i, arg in enumerate(sys.argv):
        if arg == '--phase' and i + 1 < len(sys.argv):
            phase_filter = sys.argv[i + 1]
        if arg == '--grade' and i + 1 < len(sys.argv):
            grade_filter = sys.argv[i + 1]
        if arg == '--limit' and i + 1 < len(sys.argv):
            try:
                limit = int(sys.argv[i + 1])
            except ValueError:
                pass

    api_key = os.environ.get('ANTHROPIC_AUTH_TOKEN') or os.environ.get('ANTHROPIC_API_KEY')
    if not api_key and not list_only:
        print('ERROR: No API key. Set ANTHROPIC_AUTH_TOKEN or ANTHROPIC_API_KEY.')
        sys.exit(1)

    with open(CLASSIFY, 'r', encoding='utf8') as f:
        classify = json.load(f)

    targets = collect_targets(classify, phase_filter, grade_filter)

    # Sort: very_low first (most benefit from repair), then by rows descending
    grade_order = {"very_low": 0, "low": 1, "medium": 2}
    targets.sort(key=lambda t: (grade_order.get(t["grade"], 3), -t["rows"]))

    if limit:
        targets = targets[:limit]

    # Filter out targets with no source path
    missing = [t for t in targets if not t["src_path"] or not os.path.exists(t["src_path"])]
    targets = [t for t in targets if t["src_path"] and os.path.exists(t["src_path"])]

    # Filter out already done (unless --force)
    if not force:
        targets = [t for t in targets if not os.path.exists(t["out_path"]) or os.path.getsize(t["out_path"]) == 0]

    print(f"Repair targets: {len(targets)} (missing source: {len(missing)})")
    for phase in ["textual", "scanned", "excel"]:
        n = sum(1 for t in targets if t["phase"] == phase)
        if n:
            print(f"  {phase}: {n}")

    if list_only:
        for t in targets:
            print(f"  [{t['phase'][:4]}] [{t['grade']:8s}] {t['file'][:60]} rows={t['rows']} mapped={t['mapped_cols']} empty={t['empty_rate']:.2f}")
        if missing:
            print(f"\nMissing source ({len(missing)}):")
            for t in missing:
                print(f"  [{t['phase'][:4]}] {t['file'][:60]}")
        return

    if not targets:
        print("No targets to process.")
        return

    lines = [f"=== Vision Repair ({len(targets)} files) ==="]
    model = os.environ.get('ANTHROPIC_MODEL', 'claude-sonnet-4-6')
    base_url = os.environ.get('ANTHROPIC_BASE_URL', '(default)')
    lines.append(f"Model: {model}, Base: {base_url}")

    ok = err = 0
    total_api_calls = 0
    repair_results = []

    from tqdm import tqdm
    pbar = tqdm(targets, desc='🔧 修复合并单元格', unit='file',
                bar_format='{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]')

    for ci, tgt in enumerate(targets):
        phase = tgt["phase"]
        name = tgt["file"]
        folder = tgt["folder"]
        src_path = tgt["src_path"]
        out_path = tgt["out_path"]

        pbar.set_description(f'🔧 {name[:30]}')
        lines.append(f"\n  [{ci+1}/{len(targets)}] [{phase}] {name[:55]} (grade={tgt['grade']}, rows={tgt['rows']})")

        # ─── Excel repair: expand merged cells directly ───
        if phase == "excel":
            ext = os.path.splitext(src_path)[1].lower()
            if ext == ".xls":
                n_rows, issue = repair_xls_merged_cells(src_path, out_path)
            else:
                n_rows, issue = repair_excel_merged_cells(src_path, out_path)

            if issue:
                err += 1
                lines.append(f"    ERROR: {issue[:100]}")
                repair_results.append({"file": name, "folder": folder, "phase": phase, "status": "error", "issue": issue})
            else:
                ok += 1
                lines.append(f"    OK: {n_rows} rows (merged cells expanded)")
                repair_results.append({"file": name, "folder": folder, "phase": phase, "status": "ok", "rows": n_rows, "out": tgt["out_name"]})
            pbar.update(1)
            pbar.set_postfix_str(f'ok={ok} err={err}')
            continue

        # ─── PDF repair: convert to images and send to vision ───
        try:
            from pdf2image import convert_from_path
            images = convert_from_path(src_path, dpi=DPI)
        except Exception as e:
            err += 1
            lines.append(f"    ERROR pdf2image: {str(e)[:100]}")
            repair_results.append({"file": name, "folder": folder, "phase": phase, "status": "error", "issue": f"pdf2image: {str(e)[:200]}"})
            pbar.update(1)
            pbar.set_postfix_str(f'ok={ok} err={err}')
            continue

        file_pages_data = []
        auth_error = False

        for page_num in range(1, len(images) + 1):
            img = images[page_num - 1]
            total_api_calls += 1

            buf = io.BytesIO()
            img.save(buf, format='PNG')
            img_bytes = buf.getvalue()

            t0 = time.time()
            v_rows, v_err = vision_extract_page(img_bytes, page_num)
            elapsed = round(time.time() - t0, 1)

            if v_err:
                if '401' in v_err or '无效的令牌' in v_err or 'authentication' in v_err.lower():
                    auth_error = True
                    lines.append(f"    p{page_num}: AUTH ERROR - aborting. {v_err[:100]}")
                    break
                lines.append(f"    p{page_num}: FAIL ({elapsed}s) {v_err[:100]}")
                file_pages_data.append((page_num, [[f'Error: {v_err[:200]}']], 'error'))
            elif v_rows and len(v_rows) >= 2:
                quality = 'good' if len(v_rows) >= 3 else 'partial'
                file_pages_data.append((page_num, v_rows, quality))
                lines.append(f"    p{page_num}: OK ({elapsed}s) {len(v_rows)} rows, q={quality}")
            else:
                file_pages_data.append((page_num, v_rows or [['empty']], 'poor'))
                lines.append(f"    p{page_num}: EMPTY ({elapsed}s)")

            # Rate limiting
            time.sleep(1)

        if auth_error:
            err += 1
            repair_results.append({"file": name, "folder": folder, "phase": phase, "status": "auth_error", "issue": "API auth error"})
            if file_pages_data:
                save_repair_xlsx(file_pages_data, out_path)
                lines.append(f"    Saved partial: {tgt['out_name']}")
            print(f"\nFATAL: API auth error. Fix token and re-run.")
            pbar.update(1)
            break

        if file_pages_data:
            try:
                save_repair_xlsx(file_pages_data, out_path)
                ok += 1
                total_rows = sum(len(r) for _, r, _ in file_pages_data)
                qualities = [q for _, _, q in file_pages_data]
                worst_q = 'error' if 'error' in qualities else ('poor' if 'poor' in qualities else ('partial' if 'partial' in qualities else 'good'))
                lines.append(f"    SAVED {tgt['out_name']} ({total_rows} rows, q={worst_q})")
                repair_results.append({
                    "file": name, "folder": folder, "phase": phase, "status": "ok",
                    "out": tgt["out_name"], "rows": total_rows, "quality": worst_q,
                    "pages": len(file_pages_data),
                    "page_details": [{"page": p, "quality": q, "rows": len(r)} for p, r, q in file_pages_data],
                })
            except Exception as e:
                err += 1
                lines.append(f"    SAVE ERROR: {str(e)[:100]}")
                repair_results.append({"file": name, "folder": folder, "phase": phase, "status": "error", "issue": str(e)[:200]})
        else:
            err += 1
            repair_results.append({"file": name, "folder": folder, "phase": phase, "status": "empty", "issue": "no pages"})

        pbar.update(1)
        pbar.set_postfix_str(f'ok={ok} err={err}')

    pbar.close()

    # Summary
    lines.append(f"\n=== Summary ===")
    lines.append(f"OK: {ok}, Errors: {err}, Total: {len(targets)}")
    lines.append(f"Vision API calls: {total_api_calls}")

    # Save results to classify
    classify["vision_repair_results"] = repair_results
    classify["vision_repair_summary"] = {
        "ok": ok, "errors": err, "total": len(targets),
        "api_calls": total_api_calls, "model": model,
    }
    with open(CLASSIFY, 'w', encoding='utf8') as f:
        json.dump(classify, f, ensure_ascii=False, indent=2)

    # Write log
    with open(LOG_FILE, 'w', encoding='utf8') as f:
        f.write('\n'.join(lines))

    print(f"\nDone. OK={ok}, Errors={err}")
    print(f"Repair outputs: {REPAIR_DIR}")
    print(f"Log: {LOG_FILE}")


if __name__ == '__main__':
    main()
