"""
Phase 3: Extract tables from scanned PDFs via OCR, with vision model fallback.
Each PDF -> one xlsx in temp/extracted_scanned/, each page as a sheet.

Strategy:
  1. Convert PDF pages to images (pdf2image)
  2. OCR each page with pytesseract (chi_sim + eng)
  3. Assess quality: if table structure is poor, mark for vision fallback
  4. For poor-quality pages, call Anthropic vision API to extract table

Run: python scripts/extract_scanned_tables.py [--force] [--vision-only] [--disable-vision] [--limit N]
"""
import json, os, re, sys, logging, threading, traceback, base64, time, io

os.chdir(r'D:\Research\202603-自动化债务报告')

TEMP = 'data/celma-major-events-attachments/temp'
CLASSIFY = os.path.join(TEMP, 'classify_results.json')
OUT_DIR = os.path.join(TEMP, 'extracted_scanned')
LOG_FILE = os.path.join(TEMP, 'extract_scanned_log.txt')
TIMEOUT_PER_PAGE = 90  # seconds per page for OCR/vision pipeline
DPI = 300  # resolution for pdf2image

os.makedirs(OUT_DIR, exist_ok=True)

# ---------- helpers ----------

def clean(v):
    if v is None: return ''
    s = str(v).strip().replace('\n', ' ').replace('\r', '')
    return s

def run_with_timeout(fn, timeout):
    result = [None]
    error = [None]
    def worker():
        try:
            result[0] = fn()
        except Exception as e:
            error[0] = str(e)[:300]
    t = threading.Thread(target=worker, daemon=True)
    t.start()
    t.join(timeout)
    if t.is_alive():
        return None, f'timeout after {timeout}s'
    if error[0]:
        return None, error[0]
    return result[0], None

def safe_sheet_name(name, existing):
    name = re.sub(r'[\[\]:*?/\\]', '_', name)[:28]
    if name in existing:
        for i in range(2, 100):
            n2 = f'{name}_{i}'
            if n2 not in existing:
                return n2
    return name

# ---------- OCR extraction ----------

def ocr_page_to_table(image):
    """OCR a page image and try to parse table structure.
    Returns (rows: list[list[str]], quality: str).
    quality: 'good' | 'partial' | 'poor'
    """
    import pytesseract
    # Get TSV output for structured data
    tsv = pytesseract.image_to_data(image, lang='chi_sim+eng', output_type=pytesseract.Output.DATAFRAME)
    # Also get plain text for quick assessment
    text = pytesseract.image_to_string(image, lang='chi_sim+eng')

    # Assess: does text contain table-like keywords?
    has_header = any(kw in text for kw in ['债券编码', '债券简称', '项目名称', '用途调整', '发行金额'])
    has_numbers = bool(re.search(r'\d{7,}', text))  # bond codes are 7+ digits

    if not has_header and not has_numbers:
        return [[text[:500]]], 'poor'

    # Try to reconstruct table from TSV data
    # Group by block_num and line_num to get rows
    rows = []
    if tsv is not None and len(tsv) > 0:
        tsv = tsv.dropna(subset=['text'])
        tsv = tsv[tsv['text'].str.strip() != '']
        if len(tsv) > 0:
            # Group by block and line
            for (block, par, line), group in tsv.groupby(['block_num', 'par_num', 'line_num']):
                cells = group.sort_values('left')['text'].tolist()
                row_text = [clean(c) for c in cells]
                if row_text:
                    rows.append(row_text)

    if len(rows) < 2:
        # Fallback: split by lines
        for line in text.strip().split('\n'):
            line = line.strip()
            if line:
                # Try splitting by multiple spaces or tabs
                cells = re.split(r'\s{2,}|\t', line)
                rows.append([clean(c) for c in cells])

    # Evaluate quality
    if not rows:
        return [], 'poor'

    # Check column consistency
    col_counts = [len(r) for r in rows]
    modal_cols = max(set(col_counts), key=col_counts.count) if col_counts else 0

    if modal_cols >= 10 and has_header:
        quality = 'good'
    elif modal_cols >= 5 or has_header:
        quality = 'partial'
    else:
        quality = 'poor'

    return rows, quality

# ---------- Vision model extraction ----------

def vision_extract_table(image_bytes, page_num=1):
    """Use Anthropic Claude vision to extract table from page image.
    Returns list of rows (list of strings).
    """
    import anthropic

    api_key = os.environ.get('ANTHROPIC_AUTH_TOKEN') or os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        return [], 'no API key for vision model'

    base_url = os.environ.get('ANTHROPIC_BASE_URL', '').rstrip('/')
    kwargs = {'api_key': api_key}
    if base_url:
        kwargs['base_url'] = base_url
    client = anthropic.Anthropic(**kwargs)

    # Resize large images to avoid proxy payload limits
    from PIL import Image as PILImage
    img = PILImage.open(io.BytesIO(image_bytes))
    MAX_DIM = 2400
    if max(img.size) > MAX_DIM:
        ratio = MAX_DIM / max(img.size)
        img = img.resize((int(img.width * ratio), int(img.height * ratio)))
    buf = io.BytesIO()
    img.save(buf, 'PNG')
    image_bytes = buf.getvalue()

    b64 = base64.b64encode(image_bytes).decode('utf-8')

    prompt = """请从这张图片中提取表格数据。这是一张"地方政府债券资金用途调整表"。

请严格按照以下规则输出：
1. 每行数据用 | 分隔各列
2. 保留原始表头行结构：如果表头有多行（如"一、债券信息""二、区划信息"等分类行，以及"序号""债券编码"等字段行），请原样保留每一行，不要合并
3. 保留所有数字精度，不要四舍五入
4. 空单元格用空字符串表示（两个|之间无内容）
5. 如果有合并单元格，在合并区域的第一个位置填写内容，其余留空
6. 不要添加任何解释文字，只输出表格数据

示例格式（多行表头）：
||一、债券信息|||||||||二、区划信息|||三、调整前项目信息||||||调整原因|四、调整后项目信息|||||建设期限|预计竣工日期|备注
序号|债券编码|债券简称|债券全称|发行日期|到期日期|发行利率|发行金额|未到期金额|未使用金额|用途调整金额|市县名称|区划编码|市县名称|区划编码|项目名称|项目编码|项目领域|主管部门|项目单位|建设状态|调整原因|项目名称|项目编码|项目领域|主管部门|项目单位|建设状态|建设期限|预计竣工日期|备注
1|2400001|XX专项01|...|..."""

    try:
        vision_model = os.environ.get('ANTHROPIC_MODEL', 'claude-sonnet-4-6')
        msg = client.messages.create(
            model=vision_model,
            max_tokens=8000,
            messages=[{
                'role': 'user',
                'content': [
                    {'type': 'image', 'source': {'type': 'base64', 'media_type': 'image/png', 'data': b64}},
                    {'type': 'text', 'text': prompt}
                ]
            }]
        )
        text = msg.content[0].text
        rows = []
        for line in text.strip().split('\n'):
            line = line.strip()
            if line and '|' in line:
                cells = [c.strip() for c in line.split('|')]
                rows.append(cells)
        return rows, ''
    except Exception as e:
        return [], str(e)[:200]

# ---------- Main processing ----------

def process_one_pdf(path, use_vision_fallback=True, vision_only=False):
    """Process a single scanned PDF.
    Returns list of (page_num, rows, method, quality) tuples.
    """
    from pdf2image import convert_from_path
    import io

    images = convert_from_path(path, dpi=DPI)
    all_pages = []
    stats = {
        'vision_attempts': 0,
        'vision_success': 0,
        'vision_fail': 0,
        'vision_errors': []
    }

    for pi, img in enumerate(images):
        page_num = pi + 1

        if not vision_only:
            # Step 1: Try OCR
            rows, quality = ocr_page_to_table(img)

            if quality == 'good' or (quality == 'partial' and not use_vision_fallback):
                all_pages.append((page_num, rows, 'ocr', quality))
                continue

        # Step 2: Vision fallback for poor/partial quality
        if use_vision_fallback or vision_only:
            stats['vision_attempts'] += 1
            buf = io.BytesIO()
            img.save(buf, format='PNG')
            img_bytes = buf.getvalue()

            v_rows, v_err = vision_extract_table(img_bytes, page_num)
            if v_rows and len(v_rows) >= 2:
                stats['vision_success'] += 1
                all_pages.append((page_num, v_rows, 'vision', 'good'))
                continue
            elif v_err:
                stats['vision_fail'] += 1
                stats['vision_errors'].append({'page': page_num, 'error': v_err})
                # Vision failed, use OCR result if available
                if not vision_only and rows:
                    all_pages.append((page_num, rows, 'ocr_after_vision_fail', quality))
                else:
                    all_pages.append((page_num, [[f'Vision error: {v_err}']], 'error', 'poor'))
                continue

        # Fallback: just use whatever OCR gave us
        if rows:
            all_pages.append((page_num, rows, 'ocr', quality))
        else:
            all_pages.append((page_num, [['No data extracted']], 'error', 'poor'))

    return all_pages, stats

def save_to_xlsx(pages_data, out_path):
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    sheet_names = set()
    for page_num, rows, method, quality in pages_data:
        sn = safe_sheet_name(f'p{page_num}_{method}_{quality}', sheet_names)
        sheet_names.add(sn)
        ws = wb.create_sheet(title=sn)
        # Add metadata row
        ws.append([f'# page={page_num}, method={method}, quality={quality}'])
        for row in rows:
            ws.append(row)

    if not pages_data:
        ws = wb.create_sheet(title='empty')
        ws.append(['No tables found'])

    wb.save(out_path)

def main():
    force = '--force' in sys.argv
    vision_only = '--vision-only' in sys.argv
    disable_vision = '--disable-vision' in sys.argv
    limit = None
    for arg in sys.argv:
        if arg.startswith('--limit'):
            try: limit = int(sys.argv[sys.argv.index(arg) + 1])
            except: pass

    # Check if vision API is available
    api_key = os.environ.get('ANTHROPIC_AUTH_TOKEN') or os.environ.get('ANTHROPIC_API_KEY')
    use_vision = bool(api_key) and not disable_vision

    with open(CLASSIFY, 'r', encoding='utf8') as f:
        classify = json.load(f)

    entries = classify.get('scanned', [])
    if limit:
        entries = entries[:limit]

    lines = [f'=== Extract Scanned PDFs ({len(entries)} files) ===']
    lines.append(f'Vision API: {"available" if use_vision else "NOT available"}')
    lines.append(f'Mode: {"vision-only" if vision_only else "OCR + vision fallback" if use_vision else "OCR only"}')

    results = []
    ok = skip = err = 0
    vision_calls = 0

    for i, entry in enumerate(entries):
        path = entry['path']
        name = entry['name']
        folder = os.path.basename(entry['folder'])

        out_name = f'{folder}__{os.path.splitext(name)[0]}.xlsx'
        out_name = re.sub(r'[<>:"|?*]', '_', out_name)
        out_path = os.path.join(OUT_DIR, out_name)

        if os.path.exists(out_path) and not force:
            sz = os.path.getsize(out_path)
            if sz > 0:
                skip += 1
                results.append({
                    'file': name, 'folder': folder,
                    'status': 'skipped', 'out': out_name,
                    'pages': -1, 'quality': '', 'method': '', 'issue': ''
                })
                lines.append(f'  [{i+1}/{len(entries)}] SKIP {name[:60]}')
                continue

        # Count pages first
        try:
            import fitz
            doc = fitz.open(path)
            n_pages = len(doc)
            doc.close()
        except:
            n_pages = 1

        timeout_per_page = 180 if (use_vision or vision_only) else TIMEOUT_PER_PAGE
        timeout = max(timeout_per_page * n_pages, 180)  # at least 3 minutes

        def do_extract():
            return process_one_pdf(path, use_vision_fallback=use_vision, vision_only=vision_only)

        extract_result, timeout_err = run_with_timeout(do_extract, timeout)

        if timeout_err:
            err += 1
            results.append({
                'file': name, 'folder': folder,
                'status': 'error', 'out': '',
                'pages': n_pages, 'quality': '', 'method': '', 'issue': timeout_err
            })
            lines.append(f'  [{i+1}/{len(entries)}] ERR  {name[:60]} | {timeout_err}')
            continue

        if not extract_result:
            err += 1
            results.append({
                'file': name, 'folder': folder,
                'status': 'empty', 'out': '',
                'pages': n_pages, 'quality': '', 'method': '', 'issue': 'no data'
            })
            lines.append(f'  [{i+1}/{len(entries)}] EMPTY {name[:60]}')
            continue

        pages_data, page_stats = extract_result

        # Auto-disable vision for remaining files if auth fails (401/token expired)
        if use_vision and page_stats.get('vision_fail', 0) > 0:
            errs = page_stats.get('vision_errors', [])
            auth_err = any('401' in (x.get('error', '')) or 'token' in (x.get('error', '').lower()) or '令牌' in (x.get('error', '')) for x in errs)
            if auth_err:
                use_vision = False
                lines.append('  [INFO] Vision fallback disabled for remaining files due to auth error (401/token).')

        # Summarize quality
        methods = [m for _, _, m, _ in pages_data]
        qualities = [q for _, _, _, q in pages_data]
        n_vision = sum(1 for m in methods if m == 'vision')
        vision_calls += n_vision
        worst_q = 'poor' if 'poor' in qualities else ('partial' if 'partial' in qualities else 'good')
        total_rows = sum(len(rows) for _, rows, _, _ in pages_data)

        try:
            save_to_xlsx(pages_data, out_path)
            ok += 1
            results.append({
                'file': name, 'folder': folder,
                'status': 'ok', 'out': out_name,
                'pages': n_pages, 'quality': worst_q,
                'method': 'vision' if n_vision > 0 else 'ocr',
                'total_rows': total_rows,
                'vision_pages': n_vision,
                'vision_attempts': page_stats.get('vision_attempts', 0),
                'vision_fail': page_stats.get('vision_fail', 0),
                'vision_errors': page_stats.get('vision_errors', [])[:3],
                'issue': ''
            })
            method_str = f'ocr+vision({n_vision}/{page_stats.get("vision_attempts",0)})' if page_stats.get('vision_attempts', 0) > 0 else 'ocr'
            lines.append(f'  [{i+1}/{len(entries)}] OK   {name[:50]} | {n_pages}p, {total_rows}rows, {method_str}, q={worst_q}')
            if page_stats.get('vision_fail', 0) > 0:
                first_err = page_stats.get('vision_errors', [{}])[0].get('error', '')[:120]
                lines.append(f'             vision_fail={page_stats.get("vision_fail",0)} first_error={first_err}')
        except Exception as e:
            err += 1
            results.append({
                'file': name, 'folder': folder,
                'status': 'save_error', 'out': '',
                'pages': n_pages, 'quality': '', 'method': '', 'issue': str(e)[:120]
            })
            lines.append(f'  [{i+1}/{len(entries)}] SAVE_ERR {name[:60]}')

    lines.append(f'\n=== Summary ===')
    lines.append(f'OK: {ok}, Skipped: {skip}, Errors: {err}, Total: {len(entries)}')
    lines.append(f'Vision API calls (pages): {vision_calls}')

    # Quality breakdown
    good = sum(1 for r in results if r.get('quality') == 'good')
    partial = sum(1 for r in results if r.get('quality') == 'partial')
    poor = sum(1 for r in results if r.get('quality') == 'poor')
    lines.append(f'Quality: good={good}, partial={partial}, poor={poor}')

    # Update classify_results.json
    classify['scanned_extract_results'] = results
    classify['scanned_extract_summary'] = {
        'ok': ok, 'skipped': skip, 'errors': err, 'total': len(entries),
        'vision_calls': vision_calls
    }
    with open(CLASSIFY, 'w', encoding='utf8') as f:
        json.dump(classify, f, ensure_ascii=False, indent=2)
    lines.append('Updated classify_results.json')

    with open(LOG_FILE, 'w', encoding='utf8') as f:
        f.write('\n'.join(lines))
    print(f'Done. OK={ok} Skip={skip} Err={err} Vision={vision_calls}. See {LOG_FILE}')

if __name__ == '__main__':
    main()
