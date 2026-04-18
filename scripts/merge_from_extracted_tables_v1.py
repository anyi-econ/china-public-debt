"""
Merge standard fund-usage tables from extracted artifacts.

Input sources:
  1) Original excel attachments (phase 1)
  2) temp/extracted_textual/*.xlsx (phase 2 extracted results)
  3) temp/extracted_scanned/*.xlsx (phase 3 extracted results, optional)

This script decouples extraction from merge for faster iteration.

Run examples:
  python scripts/merge_from_extracted_tables.py
  python scripts/merge_from_extracted_tables.py --include-scanned
  python scripts/merge_from_extracted_tables.py --out data/celma-major-events-attachments/merged_用途调整表.xlsx
"""

import json
import os
import re
import sys
from collections import Counter

os.chdir(r"D:\Research\202603-自动化债务报告")

TEMP = "data/celma-major-events-attachments/temp"
CLASSIFY = os.path.join(TEMP, "classify_results.json")
TEXT_DIR = os.path.join(TEMP, "extracted_textual")
SCAN_DIR = os.path.join(TEMP, "extracted_scanned")
VISION_DIR = os.path.join(TEMP, "extracted_scanned_vision")
LOG_FILE = os.path.join(TEMP, "merge_from_extracted_log.txt")

STD_COLS = [
    "序号", "债券编码", "债券简称", "债券全称", "发行日期", "到期日期",
    "发行利率", "发行金额", "未到期金额", "未使用金额", "用途调整金额",
    "市县名称（调整前）", "区划编码（调整前）", "市县名称（调整后）", "区划编码（调整后）",
    "调整前项目名称", "调整前项目编码", "调整前项目领域", "调整前主管部门",
    "调整前项目单位", "调整前建设状态", "调整原因",
    "调整后项目名称", "调整后项目编码", "调整后项目领域", "调整后主管部门",
    "调整后项目单位", "调整后建设状态", "建设期限", "预计竣工日期", "备注",
]

SKIP_KW = {"合计", "全区合计", "全省合计", "全市合计", "总计", "小计"}
HDR_LABELS = {
    "一、债券信息", "二、区划信息", "二、区域信息", "三、调整前项目信息", "四、调整后项目信息",
    "三、变动前项目信息", "四、变动后项目信息", "债券编码", "债券编 码", "发行年度",
    "债券简称", "债券全称", "发行日期", "到期日期", "发行利率", "发行金额", "未到期金额",
    "未使用金额", "用途调整金额", "序号", "项目领域", "项目名称", "项目编码", "主管部门",
    "项目单位", "建设状态", "调整原因", "建设期限", "预计竣工日期", "备注", "市县名称", "区划编码",
    "调整前项目名称", "调整后项目名称", "调整前项目编码", "调整后项目编码",
}
HDR_SET = {x.replace(" ", "") for x in HDR_LABELS}


def clean(v):
    if v is None:
        return ""
    s = str(v).strip().replace("\n", " ").replace("\r", "")
    if re.match(r"^\d+\.0$", s):
        s = s[:-2]
    return s


def norm_row(row):
    cells = [clean(c) for c in row]
    if len(cells) < 31:
        cells.extend([""] * (31 - len(cells)))
    elif len(cells) > 31:
        cells = cells[:31]
    return cells


def has_year_cols(rows):
    for row in rows[:6]:
        if "需求申报年度" in " ".join(clean(c) for c in row):
            return True
    return False


def remap_shanxi(row):
    cells = [clean(c) for c in row]
    if len(cells) < 32:
        cells.extend([""] * (32 - len(cells)))
    out = cells[:15] + cells[16:23] + cells[24:32]
    if len(out) < 31:
        out.extend([""] * (31 - len(out)))
    return out[:31]


def find_hdr(rows, kw="债券编码", mx=15):
    """Find the field-name header row (containing 债券编码).
    For multi-row headers, category rows (一、债券信息 etc.) come before
    the field-name row (序号|债券编码|...). We return the field-name row index
    so that data rows start at hi+1. The is_header_like filter in the
    data loop will also skip any remaining header-like rows below.
    """
    for i, row in enumerate(rows[:mx]):
        txt = " ".join(clean(c) for c in row)
        if kw in txt:
            return i
    return -1


def is_skip_raw_row(row):
    first = " ".join(clean(c) for c in row[:4])
    if any(kw in first for kw in SKIP_KW):
        return True
    return sum(1 for c in row if clean(c)) <= 1


def is_header_like(cells):
    seq = cells[0].replace(" ", "")
    bc = cells[1].replace(" ", "")
    if re.match(r"^[一二三四五六七八九十]、", seq):
        return True
    if bc in HDR_SET:
        return True
    non_empty = [c for c in cells if c]
    if non_empty and all(c.replace(" ", "") in HDR_SET for c in non_empty):
        return True
    if not bc and (not seq or not re.match(r"^\d", seq)):
        has_numeric = False
        for c in cells[1:]:
            v = c.replace(" ", "").replace(",", "").replace("，", "")
            if re.match(r"^\d{5,}$", v):
                has_numeric = True
                break
            if re.match(r"^\d+\.?\d*$", v) and len(v) >= 4:
                has_numeric = True
                break
        if not has_numeric:
            return True
    return False


def extract_standard_rows(rows, folder_name, file_name):
    if not rows:
        return [], "empty"

    hi = find_hdr(rows)
    if hi < 0:
        return [], "no_header"

    ncols = max((len(r) for r in rows), default=0)
    scan = rows[hi:hi + 2]
    has_bc = any("债券编码" in clean(c) for r in scan for c in r)
    has_pc = any("项目编码" in clean(c) for r in scan for c in r)

    # Keep compatibility with known "standard-like but missing 项目编码" files.
    if ncols < 20 or (not has_bc):
        return [], f"simplified_{ncols}cols"

    is_sx = has_year_cols(rows)
    out = []
    for i in range(hi + 1, len(rows)):
        row = rows[i]
        if is_skip_raw_row(row):
            continue
        cells = remap_shanxi(row) if is_sx else norm_row(row)
        if is_header_like(cells):
            continue
        # Basic data existence check
        if not cells[1] and not cells[10] and not cells[15] and not cells[22]:
            continue
        out.append([folder_name, file_name] + cells)

    if not out:
        return [], "standard_empty"
    if has_pc:
        return out, "standard"
    return out, "standard_weak"  # For special cases lacking 项目编码


def load_excel_rows(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        import xlrd

        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        return [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]

    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def merge_excel_phase(classify, lines):
    rows_all = []
    stats = Counter()
    details = []

    entries = classify.get("excel", [])
    lines.append(f"=== Merge Excel Source ({len(entries)} files) ===")
    for e in entries:
        folder_name = os.path.basename(e["folder"])
        file_name = e["name"]
        try:
            rows = load_excel_rows(e["path"])
            out, fmt = extract_standard_rows(rows, folder_name, file_name)
            stats[fmt] += 1
            rows_all.extend(out)
            details.append({"folder": folder_name, "file": file_name, "format": fmt, "rows": len(out)})
        except Exception as ex:
            stats["error"] += 1
            details.append({"folder": folder_name, "file": file_name, "format": "error", "rows": 0, "issue": str(ex)[:120]})

    lines.append(f"Excel merged rows: {len(rows_all)}")
    lines.append(f"Excel stats: {dict(stats)}")
    return rows_all, details, dict(stats)


def merge_extracted_phase(classify, key_results, base_dir, phase_name, lines):
    rows_all = []
    stats = Counter()
    details = []

    from openpyxl import load_workbook

    entries = classify.get(key_results, [])
    lines.append(f"\n=== Merge {phase_name} Extracted ({len(entries)} files) ===")

    for e in entries:
        status = e.get("status", "")
        if status not in {"ok", "skipped"}:
            stats["skip_not_ok"] += 1
            continue

        out_name = e.get("out", "")
        if not out_name:
            stats["skip_no_out"] += 1
            continue

        xlsx_path = os.path.join(base_dir, out_name)
        if not os.path.exists(xlsx_path):
            stats["missing_file"] += 1
            details.append({"folder": e.get("folder", ""), "file": e.get("file", ""), "format": "missing_file", "rows": 0})
            continue

        folder_name = e.get("folder", "")
        file_name = e.get("file", "")

        try:
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)
            file_rows = []
            file_stats = Counter()

            for ws in wb.worksheets:
                sheet_rows = [list(r) for r in ws.iter_rows(values_only=True)]
                # Remove metadata row for scanned extraction sheets
                if sheet_rows and sheet_rows[0] and str(sheet_rows[0][0] or "").startswith("# page="):
                    sheet_rows = sheet_rows[1:]

                out, fmt = extract_standard_rows(sheet_rows, folder_name, file_name)
                file_stats[fmt] += 1
                file_rows.extend(out)

            wb.close()
            rows_all.extend(file_rows)
            stats["ok"] += 1
            stats["rows"] += len(file_rows)
            details.append({
                "folder": folder_name,
                "file": file_name,
                "format": "standard" if file_rows else "non_standard",
                "rows": len(file_rows),
                "sheet_stats": dict(file_stats),
            })
        except Exception as ex:
            stats["error"] += 1
            details.append({"folder": folder_name, "file": file_name, "format": "error", "rows": 0, "issue": str(ex)[:120]})

    lines.append(f"{phase_name} merged rows: {stats.get('rows', 0)}")
    lines.append(f"{phase_name} stats: {dict(stats)}")
    return rows_all, details, dict(stats)


def merge_scanned_phase(classify, lines):
    """Merge scanned PDFs: prefer vision result from extracted_scanned_vision/
    over OCR result from extracted_scanned/."""
    rows_all = []
    stats = Counter()
    details = []

    from openpyxl import load_workbook

    ocr_entries = classify.get("scanned_ocr_results", [])
    # Build vision lookup
    vision_lookup = {}
    for vr in classify.get("scanned_vision_results", []):
        vk = vr.get("file", "") + "|" + vr.get("folder", "")
        if vr.get("status") == "ok":
            vision_lookup[vk] = vr

    lines.append(f"\n=== Merge Scanned ({len(ocr_entries)} OCR files, {len(vision_lookup)} vision overrides) ===")

    for e in ocr_entries:
        status = e.get("status", "")
        if status not in ("ok", "skipped"):
            stats["skip_not_ok"] += 1
            continue

        out_name = e.get("out", "")
        if not out_name:
            stats["skip_no_out"] += 1
            continue

        folder_name = e.get("folder", "")
        file_name = e.get("file", "")
        key = file_name + "|" + folder_name

        # Prefer vision result if available
        vr = vision_lookup.get(key)
        if vr and vr.get("out"):
            xlsx_path = os.path.join(VISION_DIR, vr["out"])
            source = "vision"
        else:
            xlsx_path = os.path.join(SCAN_DIR, out_name)
            source = "ocr"

        if not os.path.exists(xlsx_path):
            stats["missing_file"] += 1
            details.append({"folder": folder_name, "file": file_name, "format": "missing_file", "source": source, "rows": 0})
            continue

        try:
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)
            file_rows = []
            file_stats = Counter()
            for ws in wb.worksheets:
                sheet_rows = [list(r) for r in ws.iter_rows(values_only=True)]
                if sheet_rows and sheet_rows[0] and str(sheet_rows[0][0] or "").startswith("# page="):
                    sheet_rows = sheet_rows[1:]
                out, fmt = extract_standard_rows(sheet_rows, folder_name, file_name)
                file_stats[fmt] += 1
                file_rows.extend(out)
            wb.close()
            rows_all.extend(file_rows)
            stats["ok"] += 1
            stats["rows"] += len(file_rows)
            stats[f"source_{source}"] += 1
            details.append({
                "folder": folder_name, "file": file_name,
                "format": "standard" if file_rows else "non_standard",
                "source": source, "rows": len(file_rows),
                "sheet_stats": dict(file_stats),
            })
        except Exception as ex:
            stats["error"] += 1
            details.append({"folder": folder_name, "file": file_name, "format": "error", "source": source, "rows": 0, "issue": str(ex)[:120]})

    lines.append(f"Scanned merged rows: {stats.get('rows', 0)}")
    lines.append(f"Scanned stats: {dict(stats)}")
    return rows_all, details, dict(stats)


def dedupe_rows(rows):
    seen = set()
    out = []
    for r in rows:
        key = tuple(r)
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def save_output(rows, out_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "资金用途调整表"
    ws.append(["文件夹名称", "表格文件名称"] + STD_COLS)
    for r in rows:
        ws.append(r)
    wb.save(out_path)


def parse_args():
    include_scanned = "--include-scanned" in sys.argv
    out = "data/celma-major-events-attachments/merged_用途调整表.xlsx"
    if "--out" in sys.argv:
        i = sys.argv.index("--out")
        if i + 1 < len(sys.argv):
            out = sys.argv[i + 1]
    return include_scanned, out


def main():
    include_scanned, out_path = parse_args()

    with open(CLASSIFY, "r", encoding="utf8") as f:
        classify = json.load(f)

    lines = ["=== Merge From Extracted Tables ===", f"include_scanned={include_scanned}"]

    excel_rows, excel_details, excel_stats = merge_excel_phase(classify, lines)
    text_rows, text_details, text_stats = merge_extracted_phase(
        classify, "textual_extract_results", TEXT_DIR, "Textual PDF", lines
    )

    scan_rows = []
    scan_details = []
    scan_stats = {}
    if include_scanned:
        scan_rows, scan_details, scan_stats = merge_scanned_phase(
            classify, lines
        )

    all_rows = excel_rows + text_rows + scan_rows
    before = len(all_rows)
    all_rows = dedupe_rows(all_rows)
    deduped = before - len(all_rows)

    save_output(all_rows, out_path)

    # Persist summary to classify file for iterative debugging.
    classify["merge_from_extracted_summary"] = {
        "include_scanned": include_scanned,
        "excel_rows": len(excel_rows),
        "textual_rows": len(text_rows),
        "scanned_rows": len(scan_rows),
        "total_before_dedupe": before,
        "deduped_rows": deduped,
        "total_rows": len(all_rows),
        "output": out_path,
    }
    classify["merge_from_extracted_details"] = {
        "excel": {"stats": excel_stats, "files": excel_details},
        "textual": {"stats": text_stats, "files": text_details},
        "scanned": {"stats": scan_stats, "files": scan_details},
    }
    with open(CLASSIFY, "w", encoding="utf8") as f:
        json.dump(classify, f, ensure_ascii=False, indent=2)

    lines.append("\n=== Summary ===")
    lines.append(f"Output: {out_path}")
    lines.append(f"Rows before dedupe: {before}")
    lines.append(f"Deduped rows: {deduped}")
    lines.append(f"Rows final: {len(all_rows)}")
    lines.append("Updated classify_results.json")

    with open(LOG_FILE, "w", encoding="utf8") as f:
        f.write("\n".join(lines))

    print(f"Done. rows={len(all_rows)} deduped={deduped}. See {LOG_FILE}")


if __name__ == "__main__":
    main()
