"""
Phase 3b: Vision supplement for poor-quality scanned PDF pages.

Reads classify_results.json (written by extract_scanned_ocr.py) to find files
where OCR quality is poor/partial, then calls vision model page-by-page.

Each file result is saved IMMEDIATELY to extracted_scanned_vision/ after
processing, so progress is never lost. classify_results.json is also updated
after EVERY file.

Run:
  python scripts/vision_supplement_scanned.py              # process all that need it
  python scripts/vision_supplement_scanned.py --force      # re-process even if vision result exists
  python scripts/vision_supplement_scanned.py --limit 5    # cap number of files
  python scripts/vision_supplement_scanned.py --list       # just print files that need vision, don't process
  python scripts/vision_supplement_scanned.py --only-poor  # skip partial, only process poor quality

Env vars:
  ANTHROPIC_AUTH_TOKEN or ANTHROPIC_API_KEY  - API key
  ANTHROPIC_BASE_URL                        - base URL (e.g. https://yxai.anthropic.edu.pl)
  ANTHROPIC_MODEL                           - model name (default: claude-sonnet-4-6)
"""
import json, os, re, sys, base64, io, time
from tqdm import tqdm

os.chdir(r'D:\Research\202603-自动化债务报告')

TEMP = 'data/celma-major-events-attachments/temp'
CLASSIFY = os.path.join(TEMP, 'classify_results.json')
OCR_DIR = os.path.join(TEMP, 'extracted_scanned')
VISION_DIR = os.path.join(TEMP, 'extracted_scanned_vision')
LOG_FILE = os.path.join(TEMP, 'vision_supplement_log.txt')
DPI = 300
MAX_IMG_DIM = 2400  # resize images larger than this

os.makedirs(VISION_DIR, exist_ok=True)

# ---------- helpers ----------

def clean(v):
    if v is None: return ''
    return str(v).strip().replace('\n', ' ').replace('\r', '')

def safe_sheet_name(name, existing):
    name = re.sub(r'[\[\]:*?/\\]', '_', name)[:28]
    if name in existing:
        for i in range(2, 100):
            n2 = f'{name}_{i}'
            if n2 not in existing: return n2
    return name

# ---------- vision call ----------

def vision_extract_page(image_bytes, page_num=1):
    """Call vision model for one page image.
    Returns (rows: list[list[str]], error: str).
    """
    import anthropic

    api_key = os.environ.get('ANTHROPIC_AUTH_TOKEN') or os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        return [], 'no API key'

    base_url = os.environ.get('ANTHROPIC_BASE_URL', '').rstrip('/')
    kwargs = {'api_key': api_key}
    if base_url:
        kwargs['base_url'] = base_url
    client = anthropic.Anthropic(**kwargs)

    # Resize and compress to stay under 4.5MB
    from PIL import Image as PILImage
    img = PILImage.open(io.BytesIO(image_bytes))
    if max(img.size) > MAX_IMG_DIM:
        ratio = MAX_IMG_DIM / max(img.size)
        img = img.resize((int(img.width * ratio), int(img.height * ratio)))
    # Try PNG first; fall back to JPEG if too large
    buf = io.BytesIO()
    img.save(buf, 'PNG')
    image_bytes = buf.getvalue()
    media_type = 'image/png'
    if len(image_bytes) > 4_500_000:
        buf = io.BytesIO()
        img.convert('RGB').save(buf, 'JPEG', quality=85)
        image_bytes = buf.getvalue()
        media_type = 'image/jpeg'
        if len(image_bytes) > 4_500_000:
            buf = io.BytesIO()
            img.convert('RGB').save(buf, 'JPEG', quality=60)
            image_bytes = buf.getvalue()

    b64 = base64.b64encode(image_bytes).decode('utf-8')

    prompt = """请从这张图片中提取表格数据。这是一张"地方政府债券资金用途调整表"。

请严格按照以下规则输出：
1. 每行数据用 | 分隔各列
2. 保留原始表头行结构：如果表头有多行（如"一、债券信息""二、区划信息"等分类行，以及"序号""债券编码"等字段行），请原样保留每一行，不要合并
3. 保留所有数字精度，不要四舍五入
4. 空单元格用空字符串表示（两个|之间无内容）
5. 如果有合并单元格，在合并区域的第一个位置填写内容，其余留空
6. 不要添加任何解释文字，只输出表格数据

示例格式（多行表头）：
||一、债券信息|||||||||二、区划信息|||三、调整前项目信息||||||调整原因|四、调整后项目信息|||||建设期限|预计竣工日期|备注
序号|债券编码|债券简称|债券全称|发行日期|到期日期|发行利率|发行金额|未到期金额|未使用金额|用途调整金额|市县名称|区划编码|市县名称|区划编码|项目名称|项目编码|项目领域|主管部门|项目单位|建设状态|调整原因|项目名称|项目编码|项目领域|主管部门|项目单位|建设状态|建设期限|预计竣工日期|备注
1|2400001|XX专项01|...|..."""

    model = os.environ.get('ANTHROPIC_MODEL', 'claude-sonnet-4-6')
    try:
        msg = client.messages.create(
            model=model,
            max_tokens=8000,
            messages=[{
                'role': 'user',
                'content': [
                    {'type': 'image', 'source': {'type': 'base64', 'media_type': media_type, 'data': b64}},
                    {'type': 'text', 'text': prompt}
                ]
            }]
        )
        text = msg.content[0].text
        rows = []
        for line in text.strip().split('\n'):
            line = line.strip()
            if line and '|' in line:
                cells = [c.strip() for c in line.split('|')]
                rows.append(cells)
        return rows, ''
    except Exception as e:
        return [], str(e)[:300]

# ---------- save one file ----------

def save_vision_xlsx(pages_data, out_path):
    """Save vision results for one PDF.
    pages_data: list of (page_num, rows, quality_str)
    """
    from openpyxl import Workbook
    wb = Workbook(); wb.remove(wb.active)
    sheet_names = set()
    for page_num, rows, quality in pages_data:
        sn = safe_sheet_name(f'p{page_num}_vision_{quality}', sheet_names)
        sheet_names.add(sn)
        ws = wb.create_sheet(title=sn)
        ws.append([f'# page={page_num}, method=vision, quality={quality}'])
        for row in rows:
            ws.append(row)
    if not pages_data:
        ws = wb.create_sheet(title='empty'); ws.append(['No tables found'])
    wb.save(out_path)

# ---------- update JSON ----------

def update_classify(classify, key, vision_rec):
    """Update or add a vision result record in classify_results.json."""
    existing = classify.get('scanned_vision_results', [])
    found = False
    for i, r in enumerate(existing):
        if (r.get('file', '') + '|' + r.get('folder', '')) == key:
            existing[i] = vision_rec
            found = True
            break
    if not found:
        existing.append(vision_rec)
    classify['scanned_vision_results'] = existing
    with open(CLASSIFY, 'w', encoding='utf8') as f:
        json.dump(classify, f, ensure_ascii=False, indent=2)

# ---------- main ----------

def main():
    force = '--force' in sys.argv
    list_only = '--list' in sys.argv
    only_poor = '--only-poor' in sys.argv
    limit = None
    for i, arg in enumerate(sys.argv):
        if arg == '--limit' and i + 1 < len(sys.argv):
            try: limit = int(sys.argv[i + 1])
            except: pass

    # Check API key
    api_key = os.environ.get('ANTHROPIC_AUTH_TOKEN') or os.environ.get('ANTHROPIC_API_KEY')
    if not api_key and not list_only:
        print('ERROR: No API key. Set ANTHROPIC_AUTH_TOKEN or ANTHROPIC_API_KEY.')
        sys.exit(1)

    with open(CLASSIFY, 'r', encoding='utf8') as f:
        classify = json.load(f)

    # Gather files that need vision from OCR results
    ocr_results = classify.get('scanned_ocr_results', [])
    if not ocr_results:
        print('No OCR results found. Run extract_scanned_ocr.py first.')
        sys.exit(1)

    # Also build path lookup from scanned entries
    path_lookup = {}
    for entry in classify.get('scanned', []):
        k = entry['name'] + '|' + os.path.basename(entry['folder'])
        path_lookup[k] = entry['path']

    # Filter to files needing vision
    candidates = []
    for r in ocr_results:
        key = r.get('file', '') + '|' + r.get('folder', '')
        q = r.get('ocr_quality', '')
        needs = r.get('needs_vision', False)

        if only_poor and q != 'poor':
            continue
        if not needs and not force:
            continue

        out_name = re.sub(r'[<>:"|?*]', '_',
                          f'{r["folder"]}__{os.path.splitext(r["file"])[0]}.xlsx')
        vision_path = os.path.join(VISION_DIR, out_name)

        # Skip if already done (unless --force)
        if not force and os.path.exists(vision_path) and os.path.getsize(vision_path) > 0:
            continue

        pdf_path = path_lookup.get(key, '')
        if not pdf_path or not os.path.exists(pdf_path):
            continue

        candidates.append({
            'file': r['file'], 'folder': r['folder'],
            'key': key, 'path': pdf_path,
            'ocr_quality': q, 'pages': r.get('pages', 1),
            'out_name': out_name, 'vision_path': vision_path,
            'page_qualities': r.get('page_qualities', [])
        })

    if limit:
        candidates = candidates[:limit]

    # List mode
    if list_only:
        print(f'Files needing vision: {len(candidates)}')
        for c in candidates:
            pqs = ', '.join(f'p{p["page"]}={p["quality"]}' for p in c['page_qualities'])
            print(f'  [{c["ocr_quality"]}] {c["file"][:60]} | {c["pages"]}p | {pqs}')
        return

    lines = [f'=== Vision Supplement ({len(candidates)} files) ===']
    model = os.environ.get('ANTHROPIC_MODEL', 'claude-sonnet-4-6')
    base_url = os.environ.get('ANTHROPIC_BASE_URL', '(default)')
    lines.append(f'Model: {model}, Base: {base_url}')

    ok = err = 0
    total_pages_processed = 0
    total_vision_calls = 0

    pbar = tqdm(candidates, desc='👁️ 视觉补充', unit='file',
                bar_format='{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]')
    for ci, cand in enumerate(candidates):
        name = cand['file']
        folder = cand['folder']
        key = cand['key']
        pdf_path = cand['path']
        n_pages = cand['pages']
        page_qualities = cand['page_qualities']

        lines.append(f'\n  [{ci+1}/{len(candidates)}] {name[:55]} ({n_pages}p, ocr_q={cand["ocr_quality"]})')

        # Determine which pages need vision (poor/partial only)
        pages_needing_vision = []
        for pq in page_qualities:
            if pq['quality'] in ('poor', 'partial'):
                pages_needing_vision.append(pq['page'])
        if not pages_needing_vision:
            pages_needing_vision = list(range(1, n_pages + 1))

        # Convert PDF pages to images
        try:
            from pdf2image import convert_from_path
            images = convert_from_path(pdf_path, dpi=DPI)
        except Exception as e:
            err += 1
            rec = {'file': name, 'folder': folder, 'status': 'error',
                   'issue': f'pdf2image: {str(e)[:200]}',
                   'vision_pages': 0, 'vision_quality': ''}
            update_classify(classify, key, rec)
            lines.append(f'    ERROR pdf2image: {str(e)[:100]}')
            pbar.update(1)
            pbar.set_postfix_str(f'ok={ok} err={err}')
            continue

        # Process pages
        file_pages_data = []
        file_ok = True
        auth_error = False

        for page_num in pages_needing_vision:
            if page_num > len(images):
                continue

            img = images[page_num - 1]
            total_vision_calls += 1

            # Convert to bytes
            buf = io.BytesIO()
            img.save(buf, format='PNG')
            img_bytes = buf.getvalue()

            t0 = time.time()
            v_rows, v_err = vision_extract_page(img_bytes, page_num)
            elapsed = round(time.time() - t0, 1)
            total_pages_processed += 1

            if v_err:
                # Check for auth error -> abort remaining
                if '401' in v_err or 'token' in v_err.lower() or '令牌' in v_err:
                    auth_error = True
                    lines.append(f'    p{page_num}: AUTH ERROR - aborting. {v_err[:100]}')
                    break
                lines.append(f'    p{page_num}: FAIL ({elapsed}s) {v_err[:100]}')
                file_pages_data.append((page_num, [[f'Error: {v_err[:200]}']], 'error'))
            elif v_rows and len(v_rows) >= 2:
                quality = 'good' if len(v_rows) >= 3 else 'partial'
                file_pages_data.append((page_num, v_rows, quality))
                lines.append(f'    p{page_num}: OK ({elapsed}s) {len(v_rows)} rows, q={quality}')
            else:
                file_pages_data.append((page_num, v_rows or [['empty']], 'poor'))
                lines.append(f'    p{page_num}: EMPTY ({elapsed}s)')

        if auth_error:
            err += 1
            rec = {'file': name, 'folder': folder, 'status': 'auth_error',
                   'issue': 'API token expired or invalid',
                   'vision_pages': len(file_pages_data),
                   'vision_quality': ''}
            update_classify(classify, key, rec)
            lines.append(f'    ABORTED due to auth error. Stopping.')
            # Save what we have so far (if any)
            if file_pages_data:
                save_vision_xlsx(file_pages_data, cand['vision_path'])
                lines.append(f'    Saved partial: {cand["out_name"]}')
            print(f'\nFATAL: API auth error. Fix token and re-run.')
            pbar.update(1)
            pbar.set_postfix_str(f'ok={ok} err={err} AUTH_ERR')
            break  # Stop processing further files

        # Save immediately
        if file_pages_data:
            try:
                save_vision_xlsx(file_pages_data, cand['vision_path'])
                ok += 1
                qualities = [q for _, _, q in file_pages_data]
                worst_q = 'poor' if 'poor' in qualities else ('partial' if 'partial' in qualities else 'good')
                total_rows = sum(len(r) for _, r, _ in file_pages_data)
                rec = {
                    'file': name, 'folder': folder, 'status': 'ok',
                    'out': cand['out_name'],
                    'vision_pages': len(file_pages_data),
                    'vision_quality': worst_q,
                    'total_rows': total_rows,
                    'page_details': [{'page': p, 'quality': q, 'rows': len(r)}
                                     for p, r, q in file_pages_data],
                    'issue': ''
                }
                update_classify(classify, key, rec)
                lines.append(f'    SAVED {cand["out_name"]} ({total_rows} rows, q={worst_q})')
            except Exception as e:
                err += 1
                rec = {'file': name, 'folder': folder, 'status': 'save_error',
                       'issue': str(e)[:200], 'vision_pages': 0, 'vision_quality': ''}
                update_classify(classify, key, rec)
                lines.append(f'    SAVE ERROR: {str(e)[:100]}')
        else:
            err += 1
            rec = {'file': name, 'folder': folder, 'status': 'empty',
                   'issue': 'no pages extracted', 'vision_pages': 0, 'vision_quality': ''}
            update_classify(classify, key, rec)

        pbar.update(1)
        pbar.set_postfix_str(f'ok={ok} err={err}')

    pbar.close()
    # Final summary
    lines.append(f'\n=== Summary ===')
    lines.append(f'OK: {ok}, Errors: {err}, Total: {len(candidates)}')
    lines.append(f'Vision API calls: {total_vision_calls}, Pages processed: {total_pages_processed}')

    # Update summary in classify
    classify['scanned_vision_summary'] = {
        'ok': ok, 'errors': err,
        'total_candidates': len(candidates),
        'vision_calls': total_vision_calls,
        'pages_processed': total_pages_processed,
        'model': model
    }

    # Compute final_source for each scanned file
    ocr_results = classify.get('scanned_ocr_results', [])
    vision_results = {(r['file'] + '|' + r['folder']): r
                      for r in classify.get('scanned_vision_results', [])}
    for r in ocr_results:
        key = r['file'] + '|' + r['folder']
        vr = vision_results.get(key)
        if vr and vr.get('status') == 'ok':
            r['final_source'] = 'vision'
            r['vision_status'] = 'ok'
            r['vision_quality'] = vr.get('vision_quality', '')
        elif vr:
            r['final_source'] = 'ocr'
            r['vision_status'] = vr.get('status', 'error')
        elif r.get('needs_vision'):
            r['final_source'] = 'ocr'
            r['vision_status'] = 'pending'
        else:
            r['final_source'] = 'ocr'
            r['vision_status'] = 'not_needed'

    classify['scanned_ocr_results'] = ocr_results
    with open(CLASSIFY, 'w', encoding='utf8') as f:
        json.dump(classify, f, ensure_ascii=False, indent=2)
    lines.append('Updated classify_results.json with final_source fields.')

    with open(LOG_FILE, 'w', encoding='utf8') as f:
        f.write('\n'.join(lines))
    print(f'Done. OK={ok} Err={err} | Vision calls: {total_vision_calls}')
    print(f'Log: {LOG_FILE}')

if __name__ == '__main__':
    main()
