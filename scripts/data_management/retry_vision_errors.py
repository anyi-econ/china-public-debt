"""
Retry vision extraction for pages that had errors (connection, timeout, image too large).

Scans extracted_scanned_vision/*.xlsx for sheets named *_error, re-processes
those specific pages via the vision API, and updates the xlsx in-place.

Run:
  python scripts/retry_vision_errors.py                    # retry all error pages
  python scripts/retry_vision_errors.py --limit 5          # cap files
  python scripts/retry_vision_errors.py --list             # just count errors, don't process
  python scripts/retry_vision_errors.py --max-retries 3    # retry failed pages up to N times

Env vars: same as vision_supplement_scanned.py
"""
import json, os, re, sys, base64, io, time
from tqdm import tqdm
from PIL import Image as PILImage
PILImage.MAX_IMAGE_PIXELS = 300_000_000  # allow large scanned pages

os.chdir(r'D:\Research\202603-自动化债务报告')

TEMP = 'data/celma-major-events-attachments/temp'
CLASSIFY = os.path.join(TEMP, 'classify_results.json')
VISION_DIR = os.path.join(TEMP, 'extracted_scanned_vision')
LOG_FILE = os.path.join(TEMP, 'retry_vision_log.txt')
DPI = 300
MAX_IMG_DIM = 2000

# ---------- helpers ----------

def clean(v):
    if v is None: return ''
    return str(v).strip().replace('\n', ' ').replace('\r', '')

def safe_sheet_name(name, existing):
    name = re.sub(r'[\[\]:*?/\\]', '_', name)[:28]
    if name in existing:
        for i in range(2, 100):
            n2 = f'{name}_{i}'
            if n2 not in existing: return n2
    return name

# ---------- vision call (same as vision_supplement_scanned.py) ----------

def vision_extract_page(image_bytes, page_num=1):
    import anthropic

    api_key = os.environ.get('ANTHROPIC_AUTH_TOKEN') or os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        return [], 'no API key'

    base_url = os.environ.get('ANTHROPIC_BASE_URL', '').rstrip('/')
    kwargs = {'api_key': api_key}
    if base_url:
        kwargs['base_url'] = base_url
    client = anthropic.Anthropic(**kwargs)

    from PIL import Image as PILImage
    img = PILImage.open(io.BytesIO(image_bytes))
    if max(img.size) > MAX_IMG_DIM:
        ratio = MAX_IMG_DIM / max(img.size)
        img = img.resize((int(img.width * ratio), int(img.height * ratio)))

    # Compress: try PNG first, fall back to JPEG with decreasing quality
    buf = io.BytesIO()
    img.save(buf, 'PNG')
    image_bytes = buf.getvalue()
    media_type = 'image/png'
    if len(image_bytes) > 4_500_000:
        rgb_img = img.convert('RGB')
        for q in [85, 60, 40, 25]:
            buf = io.BytesIO()
            rgb_img.save(buf, 'JPEG', quality=q)
            image_bytes = buf.getvalue()
            media_type = 'image/jpeg'
            if len(image_bytes) <= 4_500_000:
                break

    b64 = base64.b64encode(image_bytes).decode('utf-8')

    prompt = """请从这张图片中提取表格数据。这是一张"地方政府债券资金用途调整表"。

请严格按照以下规则输出：
1. 每行数据用 | 分隔各列
2. 保留原始表头行结构：如果表头有多行（如"一、债券信息""二、区划信息"等分类行，以及"序号""债券编码"等字段行），请原样保留每一行，不要合并
3. 保留所有数字精度，不要四舍五入
4. 空单元格用空字符串表示（两个|之间无内容）
5. 如果有合并单元格，在合并区域的第一个位置填写内容，其余留空
6. 不要添加任何解释文字，只输出表格数据

示例格式（多行表头）：
||一、债券信息|||||||||二、区划信息|||三、调整前项目信息||||||调整原因|四、调整后项目信息|||||建设期限|预计竣工日期|备注
序号|债券编码|债券简称|债券全称|发行日期|到期日期|发行利率|发行金额|未到期金额|未使用金额|用途调整金额|市县名称|区划编码|市县名称|区划编码|项目名称|项目编码|项目领域|主管部门|项目单位|建设状态|调整原因|项目名称|项目编码|项目领域|主管部门|项目单位|建设状态|建设期限|预计竣工日期|备注
1|2400001|XX专项01|...|..."""

    model = os.environ.get('ANTHROPIC_MODEL', 'claude-sonnet-4-6')
    try:
        msg = client.messages.create(
            model=model,
            max_tokens=8000,
            messages=[{
                'role': 'user',
                'content': [
                    {'type': 'image', 'source': {'type': 'base64', 'media_type': media_type, 'data': b64}},
                    {'type': 'text', 'text': prompt}
                ]
            }]
        )
        text = msg.content[0].text
        rows = []
        for line in text.strip().split('\n'):
            line = line.strip()
            if line and '|' in line:
                cells = [c.strip() for c in line.split('|')]
                rows.append(cells)
        return rows, ''
    except Exception as e:
        return [], str(e)[:300]

# ---------- scan for error pages ----------

def scan_error_pages():
    """Scan all vision xlsx files, return list of files with error pages."""
    from openpyxl import load_workbook
    results = []  # list of {xlsx_path, xlsx_name, error_pages: [page_num, ...]}

    for fname in sorted(os.listdir(VISION_DIR)):
        if not fname.endswith('.xlsx'):
            continue
        fpath = os.path.join(VISION_DIR, fname)
        wb = load_workbook(fpath, read_only=True)
        error_pages = []
        for sn in wb.sheetnames:
            if 'error' in sn.lower():
                m = re.match(r'p(\d+)', sn)
                if m:
                    error_pages.append(int(m.group(1)))
        wb.close()
        if error_pages:
            results.append({'xlsx_path': fpath, 'xlsx_name': fname, 'error_pages': sorted(set(error_pages))})

    return results

# ---------- find source PDF path ----------

def build_pdf_lookup():
    with open(CLASSIFY, 'r', encoding='utf8') as f:
        classify = json.load(f)
    lookup = {}
    for entry in classify.get('scanned', []):
        folder = os.path.basename(entry['folder'])
        name = entry['name']
        out_name = re.sub(r'[<>:"|?*]', '_', f'{folder}__{os.path.splitext(name)[0]}.xlsx')
        lookup[out_name] = entry['path']
    return lookup

# ---------- main ----------

def main():
    from openpyxl import load_workbook, Workbook

    list_only = '--list' in sys.argv
    limit = None
    max_retries = 3
    for i, arg in enumerate(sys.argv):
        if arg == '--limit' and i + 1 < len(sys.argv):
            try: limit = int(sys.argv[i + 1])
            except: pass
        if arg == '--max-retries' and i + 1 < len(sys.argv):
            try: max_retries = int(sys.argv[i + 1])
            except: pass

    api_key = os.environ.get('ANTHROPIC_AUTH_TOKEN') or os.environ.get('ANTHROPIC_API_KEY')
    if not api_key and not list_only:
        print('ERROR: No API key. Set ANTHROPIC_AUTH_TOKEN or ANTHROPIC_API_KEY.')
        sys.exit(1)

    print('Scanning for error pages...')
    error_files = scan_error_pages()
    total_error_pages = sum(len(f['error_pages']) for f in error_files)
    print(f'Found {len(error_files)} files with {total_error_pages} error pages')

    if list_only:
        for ef in error_files:
            print(f'  {ef["xlsx_name"][:60]} | error pages: {ef["error_pages"]}')
        return

    if limit:
        error_files = error_files[:limit]

    pdf_lookup = build_pdf_lookup()
    lines = [f'=== Retry Vision Errors ({len(error_files)} files, {total_error_pages} pages) ===']

    ok_pages = err_pages = fixed_files = 0
    auth_error = False

    pbar = tqdm(error_files, desc='🔄 重试 Vision', unit='file',
                bar_format='{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]')

    for ef in pbar:
        xlsx_path = ef['xlsx_path']
        xlsx_name = ef['xlsx_name']
        error_pages = ef['error_pages']

        # Find source PDF
        pdf_path = pdf_lookup.get(xlsx_name, '')
        if not pdf_path or not os.path.exists(pdf_path):
            lines.append(f'  SKIP {xlsx_name[:50]} | PDF not found')
            pbar.set_postfix_str(f'ok={ok_pages} err={err_pages}')
            continue

        # Convert PDF to images
        try:
            from pdf2image import convert_from_path
            images = convert_from_path(pdf_path, dpi=DPI)
        except Exception as e:
            lines.append(f'  ERR  {xlsx_name[:50]} | pdf2image: {str(e)[:80]}')
            pbar.set_postfix_str(f'ok={ok_pages} err={err_pages}')
            continue

        # Load existing xlsx to update in-place
        wb = load_workbook(xlsx_path)
        file_fixed = False

        for page_num in error_pages:
            if page_num > len(images):
                continue

            # Try up to max_retries times
            success = False
            for attempt in range(max_retries):
                img = images[page_num - 1]
                buf = io.BytesIO()
                img.save(buf, format='PNG')
                img_bytes = buf.getvalue()

                t0 = time.time()
                v_rows, v_err = vision_extract_page(img_bytes, page_num)
                elapsed = round(time.time() - t0, 1)

                if v_err:
                    if 'status_code: 401' in v_err or 'authentication_error' in v_err.lower():
                        auth_error = True
                        break
                    if attempt < max_retries - 1:
                        time.sleep(3)  # brief pause before retry
                        continue
                    err_pages += 1
                    lines.append(f'    p{page_num}: FAIL after {max_retries} attempts ({elapsed}s) {v_err[:80]}')
                    break

                # Success — replace error sheet with good data
                quality = 'good' if v_rows and len(v_rows) >= 3 else ('partial' if v_rows else 'poor')

                # Remove old error sheet(s) for this page
                for sn in list(wb.sheetnames):
                    if sn.startswith(f'p{page_num}_vision_error'):
                        del wb[sn]

                # Create new sheet
                sheet_names = set(wb.sheetnames)
                sn = safe_sheet_name(f'p{page_num}_vision_{quality}', sheet_names)
                ws = wb.create_sheet(title=sn)
                ws.append([f'# page={page_num}, method=vision, quality={quality}'])
                for row in (v_rows or []):
                    ws.append(row)

                ok_pages += 1
                file_fixed = True
                success = True
                lines.append(f'    p{page_num}: OK ({elapsed}s) {len(v_rows or [])} rows, q={quality}')
                break

            if auth_error:
                break

        # Save updated xlsx
        wb.save(xlsx_path)
        wb.close()

        if file_fixed:
            fixed_files += 1
        pbar.set_postfix_str(f'ok={ok_pages} err={err_pages}')

        if auth_error:
            print('\nFATAL: API auth error. Fix token and re-run.')
            lines.append('ABORTED: auth error')
            break

    pbar.close()

    # Update classify_results.json with corrected vision quality
    with open(CLASSIFY, 'r', encoding='utf8') as f:
        classify = json.load(f)

    # Re-assess quality for updated files
    vision_results = classify.get('scanned_vision_results', [])
    vr_lookup = {}
    for vr in vision_results:
        k = vr.get('file', '') + '|' + vr.get('folder', '')
        vr_lookup[k] = vr

    # Re-scan updated xlsx files to update quality
    scanned_lookup = {}
    for entry in classify.get('scanned', []):
        folder = os.path.basename(entry['folder'])
        name = entry['name']
        out_name = re.sub(r'[<>:"|?*]', '_', f'{folder}__{os.path.splitext(name)[0]}.xlsx')
        scanned_lookup[out_name] = (name, folder)

    for ef in error_files:
        xlsx_name = ef['xlsx_name']
        info = scanned_lookup.get(xlsx_name)
        if not info:
            continue
        name, folder = info
        key = name + '|' + folder

        xlsx_path = ef['xlsx_path']
        if not os.path.exists(xlsx_path):
            continue

        wb = load_workbook(xlsx_path, read_only=True)
        page_details = []
        total_rows = 0
        has_error = False
        for sn in wb.sheetnames:
            m = re.match(r'p(\d+)_vision_(\w+)', sn)
            if not m:
                continue
            pn = int(m.group(1))
            q = m.group(2)
            if q == 'error':
                has_error = True
            ws = wb[sn]
            n_rows = sum(1 for _ in ws.iter_rows()) - 1  # subtract header comment
            total_rows += max(0, n_rows)
            page_details.append({'page': pn, 'quality': q, 'rows': max(0, n_rows)})
        wb.close()

        qualities = [p['quality'] for p in page_details if p['quality'] != 'error']
        if qualities:
            worst_q = 'poor' if 'poor' in qualities else ('partial' if 'partial' in qualities else 'good')
        else:
            worst_q = 'error' if has_error else 'poor'

        rec = vr_lookup.get(key, {})
        rec.update({
            'file': name, 'folder': folder, 'status': 'ok' if not has_error else 'partial_error',
            'out': xlsx_name, 'vision_pages': len(page_details),
            'vision_quality': worst_q, 'total_rows': total_rows,
            'page_details': page_details, 'issue': ''
        })
        vr_lookup[key] = rec

    classify['scanned_vision_results'] = list(vr_lookup.values())
    with open(CLASSIFY, 'w', encoding='utf8') as f:
        json.dump(classify, f, ensure_ascii=False, indent=2)

    lines.append(f'\n=== Summary ===')
    lines.append(f'Fixed pages: {ok_pages}, Still errored: {err_pages}, Fixed files: {fixed_files}')

    with open(LOG_FILE, 'w', encoding='utf8') as f:
        f.write('\n'.join(lines))

    print(f'\nDone. Fixed pages: {ok_pages} | Still errors: {err_pages} | Fixed files: {fixed_files}/{len(error_files)}')

if __name__ == '__main__':
    main()
