"""
Phase 2: Extract tables from textual PDFs into individual xlsx files.
Each PDF -> one xlsx in temp/extracted_textual/, each page table as a sheet.
Run: python scripts/extract_pdf_tables.py [--force]
"""
import json, os, re, sys, logging, threading, traceback
from tqdm import tqdm

logging.getLogger('pdfminer').setLevel(logging.ERROR)
os.chdir(r'D:\Research\202603-自动化债务报告')

TEMP = 'data/celma-major-events-attachments/temp'
CLASSIFY = os.path.join(TEMP, 'classify_results.json')
OUT_DIR = os.path.join(TEMP, 'extracted_textual')
LOG_FILE = os.path.join(TEMP, 'extract_textual_log.txt')
TIMEOUT = 300  # seconds per PDF

os.makedirs(OUT_DIR, exist_ok=True)

def clean(v):
    if v is None: return ''
    s = str(v).strip().replace('\n', ' ').replace('\r', '')
    if re.match(r'^\d+\.0$', s): s = s[:-2]
    return s

def run_with_timeout(fn, timeout):
    result = [None]
    error = [None]
    def worker():
        try:
            result[0] = fn()
        except Exception as e:
            error[0] = str(e)[:200]
    t = threading.Thread(target=worker, daemon=True)
    t.start()
    t.join(timeout)
    if t.is_alive():
        return None, f'timeout after {timeout}s'
    if error[0]:
        return None, error[0]
    return result[0], None

def extract_one_pdf(path):
    """Extract all tables from a textual PDF using pdfplumber.
    Returns list of (page_num, table_data) where table_data is list of rows (list of strings).
    """
    import pdfplumber
    results = []  # [(page_num, [[cell, ...], ...]), ...]
    with pdfplumber.open(path) as pdf:
        for pi, page in enumerate(pdf.pages):
            try:
                tables = page.extract_tables() or []
                for ti, t in enumerate(tables):
                    if t and len(t) > 0:
                        cleaned = []
                        for row in t:
                            cleaned.append([clean(c) for c in row])
                        results.append((pi + 1, ti, cleaned))
            except Exception as e:
                results.append((pi + 1, -1, [[f'ERROR: {str(e)[:100]}']]))
    return results

def safe_sheet_name(name, existing):
    """Create safe sheet name (max 31 chars, no special chars, unique)."""
    name = re.sub(r'[\[\]:*?/\\]', '_', name)[:28]
    if name in existing:
        for i in range(2, 100):
            n2 = f'{name}_{i}'
            if n2 not in existing:
                return n2
    return name

def save_to_xlsx(tables_data, out_path):
    """Save extracted tables to xlsx. Each table becomes a sheet."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    sheet_names = set()
    for page_num, table_idx, rows in tables_data:
        sn = safe_sheet_name(f'p{page_num}_t{table_idx}', sheet_names)
        sheet_names.add(sn)
        ws = wb.create_sheet(title=sn)
        for row in rows:
            ws.append(row)

    if not tables_data:
        ws = wb.create_sheet(title='empty')
        ws.append(['No tables found'])

    wb.save(out_path)

def main():
    force = '--force' in sys.argv

    with open(CLASSIFY, 'r', encoding='utf8') as f:
        classify = json.load(f)

    entries = classify.get('textual', [])
    lines = [f'=== Extract Textual PDFs ({len(entries)} files) ===']

    results = []
    ok = skip = err = 0

    pbar = tqdm(entries, desc='📄 标准件 PDF', unit='file',
                bar_format='{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]')
    for i, entry in enumerate(entries):
        path = entry['path']
        name = entry['name']
        folder = os.path.basename(entry['folder'])

        # Output filename: folder__name.xlsx
        out_name = f'{folder}__{os.path.splitext(name)[0]}.xlsx'
        # Sanitize filename
        out_name = re.sub(r'[<>:"|?*]', '_', out_name)
        out_path = os.path.join(OUT_DIR, out_name)

        # Skip if already extracted (unless --force)
        if os.path.exists(out_path) and not force:
            sz = os.path.getsize(out_path)
            if sz > 0:
                skip += 1
                results.append({
                    'file': name, 'folder': folder,
                    'status': 'skipped', 'out': out_name,
                    'tables': -1, 'issue': ''
                })
                lines.append(f'  [{i+1}/{len(entries)}] SKIP {name[:60]} (exists)')
                pbar.update(1)
                pbar.set_postfix_str(f'ok={ok} skip={skip} err={err}')
                continue

        # Extract
        def do_extract():
            return extract_one_pdf(path)
        tables_data, timeout_err = run_with_timeout(do_extract, TIMEOUT)

        if timeout_err:
            err += 1
            results.append({
                'file': name, 'folder': folder,
                'status': 'error', 'out': '',
                'tables': 0, 'issue': timeout_err
            })
            lines.append(f'  [{i+1}/{len(entries)}] ERR  {name[:60]} | {timeout_err}')
            pbar.update(1)
            pbar.set_postfix_str(f'ok={ok} skip={skip} err={err}')
            continue

        if not tables_data:
            err += 1
            results.append({
                'file': name, 'folder': folder,
                'status': 'empty', 'out': '',
                'tables': 0, 'issue': 'no tables found'
            })
            lines.append(f'  [{i+1}/{len(entries)}] EMPTY {name[:60]}')
            pbar.update(1)
            pbar.set_postfix_str(f'ok={ok} skip={skip} err={err}')
            continue

        # Count actual data tables (exclude error markers)
        real_tables = [t for t in tables_data if t[1] >= 0]
        total_rows = sum(len(t[2]) for t in real_tables)

        # Save
        try:
            save_to_xlsx(tables_data, out_path)
            ok += 1
            results.append({
                'file': name, 'folder': folder,
                'status': 'ok', 'out': out_name,
                'tables': len(real_tables), 'total_rows': total_rows,
                'issue': ''
            })
            lines.append(f'  [{i+1}/{len(entries)}] OK   {name[:60]} | {len(real_tables)} tables, {total_rows} rows')
        except Exception as e:
            err += 1
            results.append({
                'file': name, 'folder': folder,
                'status': 'save_error', 'out': '',
                'tables': 0, 'issue': str(e)[:120]
            })
            lines.append(f'  [{i+1}/{len(entries)}] SAVE_ERR {name[:60]} | {e}')

        pbar.update(1)
        pbar.set_postfix_str(f'ok={ok} skip={skip} err={err}')

    pbar.close()
    lines.append(f'\n=== Summary ===')
    lines.append(f'OK: {ok}, Skipped: {skip}, Errors: {err}, Total: {len(entries)}')

    # Update classify_results.json
    classify['textual_extract_results'] = results
    classify['textual_extract_summary'] = {'ok': ok, 'skipped': skip, 'errors': err, 'total': len(entries)}
    with open(CLASSIFY, 'w', encoding='utf8') as f:
        json.dump(classify, f, ensure_ascii=False, indent=2)
    lines.append('Updated classify_results.json')

    # Write log
    with open(LOG_FILE, 'w', encoding='utf8') as f:
        f.write('\n'.join(lines))
    print(f'Done. OK={ok} Skip={skip} Err={err}. See {LOG_FILE}')

if __name__ == '__main__':
    main()
