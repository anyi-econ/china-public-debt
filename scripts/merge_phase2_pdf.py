"""
Phase 2: Extract and merge textual PDF-based 资金用途调整表 files.
Appends results to the existing merged_用途调整表.xlsx.
Updates classify_results.json with per-file status.
"""
import json, os, re, sys
import pdfplumber
from openpyxl import load_workbook

# Suppress pdfminer CropBox warnings
import logging
logging.getLogger('pdfminer').setLevel(logging.ERROR)

TEMP = 'data/celma-major-events-attachments/temp'
CLASSIFY = os.path.join(TEMP, 'classify_results.json')
OUT_XLSX = 'data/celma-major-events-attachments/merged_用途调整表.xlsx'

# Standard 30+1 column header
STD_COLS = [
    '序号',
    '债券编码', '债券简称', '债券全称', '发行日期', '到期日期',
    '发行利率', '发行金额', '未到期金额', '未使用金额', '用途调整金额',
    '市县名称（调整前）', '区划编码（调整前）', '市县名称（调整后）', '区划编码（调整后）',
    '调整前项目名称', '调整前项目编码', '调整前项目领域', '调整前主管部门',
    '调整前项目单位', '调整前建设状态', '调整原因',
    '调整后项目名称', '调整后项目编码', '调整后项目领域', '调整后主管部门',
    '调整后项目单位', '调整后建设状态', '建设期限', '预计竣工日期',
    '备注',
]

HEADER_KEYWORDS = {'债券编码', '债券简称', '发行日期', '用途调整', '项目名称', '项目编码'}
SKIP_KEYWORDS = {'合计', '全区合计', '全省合计', '全市合计', '总计', '小计'}


def clean_cell(v):
    if v is None:
        return ''
    s = str(v).strip().replace('\n', ' ').replace('\r', '')
    if re.match(r'^\d+\.0$', s):
        s = s[:-2]
    return s


def find_header_row_in_table(table, max_scan=5):
    """Find header row index in a pdfplumber table."""
    for i, row in enumerate(table[:max_scan]):
        text = ' '.join(clean_cell(c) for c in row)
        if '债券编码' in text:
            return i
    return -1


def is_skip_row(row):
    first_cells = ' '.join(clean_cell(c) for c in row[:4])
    for kw in SKIP_KEYWORDS:
        if kw in first_cells:
            return True
    non_empty = sum(1 for c in row if clean_cell(c))
    if non_empty <= 1:
        return True
    return False


def is_standard_table(table, ncols):
    """Check if a table is standard format based on column count and header keywords."""
    if ncols < 20:
        return False
    for row in table[:5]:
        text = ' '.join(clean_cell(c) for c in row)
        if '债券编码' in text and '项目' in text:
            return True
    return False


def normalize_row(row):
    """Normalize a standard-format row to 31 columns."""
    cells = [clean_cell(c) for c in row]
    if len(cells) < 31:
        cells.extend([''] * (31 - len(cells)))
    elif len(cells) > 31:
        cells = cells[:31]
    return cells


def has_extra_year_cols(table):
    """Check if table has 需求申报年度 columns."""
    for row in table[:5]:
        text = ' '.join(clean_cell(c) for c in row)
        if '需求申报年度' in text:
            return True
    return False


def remap_shanxi_row(row):
    """Remap rows with 需求申报年度 columns."""
    cells = [clean_cell(c) for c in row]
    if len(cells) < 32:
        cells.extend([''] * (32 - len(cells)))
    out = cells[:15] + cells[16:23] + cells[24:32]
    if len(out) < 31:
        out.extend([''] * (31 - len(out)))
    return out[:31]


def extract_tables_from_pdf(path):
    """Extract all tables from all pages of a PDF."""
    all_tables = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for t in tables:
                if t and len(t) > 0:
                    all_tables.append(t)
    return all_tables


def merge_continuation_tables(tables):
    """Merge tables that span multiple pages into one.
    If subsequent tables don't have the header row, they're continuation."""
    if not tables:
        return []

    merged = []
    current = None

    for t in tables:
        ncols = max(len(r) for r in t) if t else 0
        has_header = find_header_row_in_table(t) >= 0

        if has_header:
            if current is not None:
                merged.append(current)
            current = list(t)
        elif current is not None:
            # Continuation - append rows (skip empty/header-like rows)
            for row in t:
                # Ensure same column count
                if len(row) < len(current[0]):
                    row = list(row) + [''] * (len(current[0]) - len(row))
                elif len(row) > len(current[0]):
                    row = row[:len(current[0])]
                current.append(row)
        else:
            # No header and no current - check if it's a standard table
            if ncols >= 20:
                current = list(t)
            else:
                # Isolated small table, skip or save separately
                merged.append(t)

    if current is not None:
        merged.append(current)

    return merged


def process_pdf_file(entry):
    """Process a single textual PDF file. Returns (data_rows, classify_info)."""
    path = entry['path']
    folder = entry['folder']
    name = entry['name']
    folder_name = os.path.basename(folder)

    classify_entry = {
        'path': path,
        'folder_name': folder_name,
        'file_name': name,
        'file_type': 'textual_pdf',
        'source': 'phase2_pdf',
    }

    try:
        raw_tables = extract_tables_from_pdf(path)
    except Exception as e:
        classify_entry.update({
            'mergeable': False,
            'format': 'error',
            'issue': f'PDF读取失败: {e}',
        })
        return [], classify_entry

    if not raw_tables:
        classify_entry.update({
            'mergeable': False,
            'format': 'no_tables',
            'issue': '未提取到表格',
        })
        return [], classify_entry

    # Merge continuation tables
    tables = merge_continuation_tables(raw_tables)

    # Find the main standard-format table
    standard_table = None
    for t in tables:
        ncols = max(len(r) for r in t) if t else 0
        if is_standard_table(t, ncols):
            standard_table = t
            break

    if standard_table is None:
        # Check if any table has enough columns
        for t in tables:
            ncols = max(len(r) for r in t) if t else 0
            if ncols >= 20:
                standard_table = t
                break

    if standard_table is None:
        # Simplified or non-standard format
        # Get info about the largest table
        largest = max(tables, key=lambda t: len(t))
        ncols = max(len(r) for r in largest) if largest else 0
        sample_headers = []
        for row in largest[:3]:
            cleaned = [clean_cell(c) for c in row if clean_cell(c)]
            sample_headers.extend(cleaned)
        classify_entry.update({
            'mergeable': False,
            'format': 'simplified',
            'issue': f'简化格式 ({ncols}列), 表头: {", ".join(sample_headers[:10])[:120]}',
            'ncols': ncols,
        })
        return [], classify_entry

    # Process the standard table
    ncols = max(len(r) for r in standard_table)
    hdr_idx = find_header_row_in_table(standard_table)
    if hdr_idx < 0:
        hdr_idx = 0  # Assume first row is header-ish

    is_shanxi = has_extra_year_cols(standard_table)

    data_rows = []
    start_idx = hdr_idx + 1
    for i in range(start_idx, len(standard_table)):
        row = standard_table[i]
        if is_skip_row(row):
            continue

        if is_shanxi:
            cells = remap_shanxi_row(row)
        else:
            cells = normalize_row(row)

        # Validate: at least some key fields non-empty
        bond_code = cells[1]
        project_before = cells[15]
        project_after = cells[22]
        adjust_amount = cells[10]

        if not bond_code and not project_before and not project_after and not adjust_amount:
            continue

        data_rows.append(cells)

    if len(data_rows) == 0:
        classify_entry.update({
            'mergeable': False,
            'format': 'standard_empty',
            'issue': '标准格式但无有效数据行',
        })
        return [], classify_entry

    classify_entry.update({
        'mergeable': True,
        'format': 'standard' + ('_shanxi' if is_shanxi else ''),
        'issue': '',
        'data_rows': len(data_rows),
        'ncols': ncols,
    })
    return data_rows, classify_entry


def main():
    with open(CLASSIFY, 'r', encoding='utf8') as f:
        classify = json.load(f)

    pdf_entries = classify.get('textual', [])
    print(f'Total textual PDF files: {len(pdf_entries)}')

    file_results = []
    all_data_rows = []
    standard_count = 0
    simplified_count = 0
    error_count = 0

    for entry in pdf_entries:
        folder_name = os.path.basename(entry['folder'])
        file_name = entry['name']

        data_rows, classify_entry = process_pdf_file(entry)

        if data_rows:
            for row in data_rows:
                all_data_rows.append([folder_name, file_name] + row)
            standard_count += 1
            print(f'  [OK] {file_name} | {len(data_rows)} rows')
        else:
            fmt = classify_entry.get('format', 'unknown')
            issue = classify_entry.get('issue', '')
            if fmt == 'simplified':
                simplified_count += 1
                print(f'  [SKIP] SIMPLIFIED: {file_name} | {issue[:80]}')
            else:
                error_count += 1
                print(f'  [ERR] {fmt}: {file_name} | {issue[:80]}')

        file_results.append(classify_entry)

    print(f'\n=== PHASE 2 SUMMARY ===')
    print(f'Standard mergeable: {standard_count} files, {len(all_data_rows)} new rows')
    print(f'Simplified (deferred): {simplified_count} files')
    print(f'Errors: {error_count} files')

    # Append to existing merged file
    if all_data_rows and os.path.exists(OUT_XLSX):
        wb = load_workbook(OUT_XLSX)
        ws = wb.active
        existing_rows = ws.max_row - 1  # Subtract header
        for row in all_data_rows:
            ws.append(row)
        wb.save(OUT_XLSX)
        print(f'\nAppended {len(all_data_rows)} rows to {OUT_XLSX} (total: {existing_rows + len(all_data_rows)})')
    elif all_data_rows:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '资金用途调整表'
        ws.append(['文件夹名称', '表格文件名称'] + STD_COLS)
        for row in all_data_rows:
            ws.append(row)
        wb.save(OUT_XLSX)
        print(f'\nCreated {OUT_XLSX} with {len(all_data_rows)} rows')

    # Update classify_results.json
    classify['pdf_results'] = file_results
    classify['pdf_summary'] = {
        'total': len(pdf_entries),
        'standard_merged': standard_count,
        'simplified_deferred': simplified_count,
        'errors': error_count,
        'total_data_rows': len(all_data_rows),
    }
    with open(CLASSIFY, 'w', encoding='utf8') as f:
        json.dump(classify, f, ensure_ascii=False, indent=2)
    print(f'Updated {CLASSIFY}')


if __name__ == '__main__':
    main()
