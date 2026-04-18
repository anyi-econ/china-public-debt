"""
Merge standard fund-usage tables from extracted artifacts (v2 - semantic alignment).

Input sources:
  1) Original excel attachments (phase 1)
  2) temp/extracted_textual/*.xlsx (phase 2 extracted results)
  3) temp/extracted_scanned/*.xlsx (phase 3 extracted results, optional)

Key improvements over v1:
  - Semantic column alignment: maps source headers to STD_COLS by name, not position
  - "识别方法" first column in output (Excel表格/标准件提取/扫描件OCR/扫描件vision)
  - Multi-row header support (category + field rows)
  - Simplified format capture (maps partial columns into STD_COLS)
  - Quality assessment per file: empty-rate scoring, merged-cell detection
  - Merged cell gap-filling (inherit from above for empty cells in merged regions)

Run examples:
  python scripts/merge_from_extracted_tables.py
  python scripts/merge_from_extracted_tables.py --include-scanned
  python scripts/merge_from_extracted_tables.py --out merged_xxx.xlsx
"""

import json
import os
import re
import sys
from collections import Counter

os.chdir(r"D:\Research\202603-自动化债务报告")

TEMP = "data/celma-major-events-attachments/temp"
CLASSIFY = os.path.join(TEMP, "classify_results.json")
TEXT_DIR = os.path.join(TEMP, "extracted_textual")
SCAN_DIR = os.path.join(TEMP, "extracted_scanned")
VISION_DIR = os.path.join(TEMP, "extracted_scanned_vision")
LOG_FILE = os.path.join(TEMP, "merge_from_extracted_log.txt")

# === 31 standard columns ===
STD_COLS = [
    "序号", "债券编码", "债券简称", "债券全称", "发行日期", "到期日期",
    "发行利率", "发行金额", "未到期金额", "未使用金额", "用途调整金额",
    "市县名称（调整前）", "区划编码（调整前）", "市县名称（调整后）", "区划编码（调整后）",
    "调整前项目名称", "调整前项目编码", "调整前项目领域", "调整前主管部门",
    "调整前项目单位", "调整前建设状态", "调整原因",
    "调整后项目名称", "调整后项目编码", "调整后项目领域", "调整后主管部门",
    "调整后项目单位", "调整后建设状态", "建设期限", "预计竣工日期", "备注",
]
STD_IDX = {col: i for i, col in enumerate(STD_COLS)}

# === Alias map: variant name -> canonical STD_COLS name ===
ALIAS_MAP = {
    # Bond info variants
    "债券编 码": "债券编码", "债券\n编码": "债券编码", "债券编\n码": "债券编码",
    "债券\n简称": "债券简称", "债券简述": "债券简称", "债券 简称": "债券简称",
    "债券\n全称": "债券全称", "债券 全称": "债券全称",
    "发行\n日期": "发行日期", "发行\n利率": "发行利率",
    "发行利率（%）": "发行利率", "发行利率（票面）": "发行利率",
    "到期\n日期": "到期日期",
    "发行\n金额": "发行金额", "未到期\n金额": "未到期金额",
    "未使用\n金额": "未使用金额", "用途\n调整\n金额": "用途调整金额",
    "用途调整\n金额": "用途调整金额", "用途变动金额": "用途调整金额",
    "用途变动\n金额": "用途调整金额",
    # Region info variants
    "市县\n名称（调整前）": "市县名称（调整前）",
    "区划\n编码（调整前）": "区划编码（调整前）",
    "市县\n名称（调整后）": "市县名称（调整后）",
    "区划\n编码\n（调整后）": "区划编码（调整后）",
    "区划\n编码（调整后）": "区划编码（调整后）",
    "区划名称（调整前）": "市县名称（调整前）",
    "区划名称（调整后）": "市县名称（调整后）",
    "区划名\n称（调\n整前）": "市县名称（调整前）",
    "区划编\n码（调\n整前）": "区划编码（调整前）",
    "市县名\n称（调\n整前）": "市县名称（调整前）",
    "市县名称\n（调整\n前）": "市县名称（调整前）",
    "市县名称\n（调整\n后）": "市县名称（调整后）",
    "区划编码\n（调整\n前）": "区划编码（调整前）",
    "区划编码\n（调整\n后）": "区划编码（调整后）",
    # 变动=调整 (四川 etc.)
    "变动前": "调整前", "变动后": "调整后",
    # Simplified name variants
    "债券名称": "债券简称", "对应债券名称": "债券简称", "所属债券": "债券简称",
    "债券期限": "建设期限", "债券年限": "建设期限", "期限": "建设期限",
    "调整金额": "用途调整金额", "拟调整额度": "用途调整金额",
    "项目发行金额": "发行金额", "项目调整金额": "用途调整金额",
    "项目代码": "调整前项目编码", "金额": "用途调整金额",
    "原债券安排金额": "发行金额", "已用作资本金金额": "用途调整金额",
    # Region simplifications
    "地区": "市县名称（调整前）", "盟市": "市县名称（调整前）",
    "地市": "市县名称（调整前）", "所属市县": "市县名称（调整前）",
    "县区": "区划编码（调整前）", "市州": "市县名称（调整前）",
    "区划": "市县名称（调整前）",
    # Project info simplifications
    "项目名称": "调整前项目名称",
    "调整前项目信息": "调整前项目名称",
    "调整后项目信息": "调整后项目名称",
    "额度": "发行金额",
    "调入金额": "用途调整金额", "调出金额": "发行金额",
    "拟调出金额": "发行金额", "拟调入金额": "用途调整金额",
    "变更前项目基本情况": "调整前项目名称",
    "变更后项目基本情况": "调整后项目名称",
    "变更原因": "调整原因",
    "建设状态": "调整前建设状态",
    "建设状态（未开工/在建/已竣工）": "调整前建设状态",
    "年度": "发行日期", "发行年份": "发行日期",
    "单位名称": "调整前项目单位",
    "项目单位": "调整前项目单位", "主管部门": "调整前主管部门",
    "资金投向领域": "调整前项目领域", "项目领域": "调整前项目领域",
    "项目编码": "调整前项目编码",
    "债券类型": "备注", "项目主要建设内容和规模": "备注",
    "调整前金额": "发行金额", "调整后金额": "用途调整金额",
    "债券地方（省级）": "债券编码",
    "市县名称（清单）": "市县名称（调整前）",
    "图例类型": "备注", "变更类型": "备注",
    "项目目录类型": "调整前项目领域",
}

# Category row labels to skip
CATEGORY_LABELS = {
    "一、债券信息", "二、区划信息", "二、区域信息",
    "三、调整前项目信息", "四、调整后项目信息",
    "三、变动前项目信息", "四、变动后项目信息",
    "五、其他", "调整原因及备注",
}

SKIP_KW = {"合计", "全区合计", "全省合计", "全市合计", "总计", "小计"}

HDR_LABELS = {
    "一、债券信息", "二、区划信息", "二、区域信息",
    "三、调整前项目信息", "四、调整后项目信息",
    "三、变动前项目信息", "四、变动后项目信息",
    "债券编码", "债券编 码", "发行年度",
    "债券简称", "债券全称", "发行日期", "到期日期", "发行利率", "发行金额",
    "未到期金额", "未使用金额", "用途调整金额", "序号", "项目领域", "项目名称",
    "项目编码", "主管部门", "项目单位", "建设状态", "调整原因", "建设期限",
    "预计竣工日期", "备注", "市县名称", "区划编码",
    "调整前项目名称", "调整后项目名称", "调整前项目编码", "调整后项目编码",
    "变更类型", "变更前项目基本情况", "变更后项目基本情况",
    "用途调整前", "用途调整后", "变动前", "变动后",
    "调整前项目信息", "调整后项目信息",
    "拟调出金额", "拟调入金额", "调出项目", "调入项目",
    "项目发行金额", "项目调整金额", "对应债券名称",
    "调整前金额", "调整后金额", "额度", "拟调整额度",
    "地区", "盟市", "地市", "市州", "县区",
}
HDR_SET = {x.replace(" ", "").replace("\n", "") for x in HDR_LABELS}


# ── helpers ──────────────────────────────────────────────────────

def clean(v):
    if v is None:
        return ""
    s = str(v).strip().replace("\r", "")
    if "\n" in s:
        s = re.sub(r"\n+", "\n", s)
    if re.match(r"^\d+\.0$", s):
        s = s[:-2]
    return s


def clean_flat(v):
    """Clean and flatten all newlines to spaces."""
    s = clean(v)
    return s.replace("\n", " ").strip()


def resolve_alias(raw):
    """Resolve a raw column header to a STD_COLS name, or None."""
    s = raw.strip()
    if not s:
        return None

    # Direct match in STD_COLS
    flat = s.replace("\n", "").replace(" ", "")
    for col in STD_COLS:
        if flat == col.replace(" ", ""):
            return col

    # Alias map: try raw, then flattened variants
    if s in ALIAS_MAP:
        return ALIAS_MAP[s]
    if flat in ALIAS_MAP:
        return ALIAS_MAP[flat]
    flat_sp = re.sub(r"\s+", "", s.replace("\n", " "))
    if flat_sp in ALIAS_MAP:
        return ALIAS_MAP[flat_sp]

    # Fuzzy substring match for longer alias keys
    for alias_key, target in ALIAS_MAP.items():
        ak = alias_key.replace("\n", "").replace(" ", "")
        if ak and len(ak) >= 4 and ak in flat:
            return target

    return None


def is_category_row(row):
    non_empty = [clean(c) for c in row if clean(c)]
    if not non_empty:
        return False
    cat_set = {c.replace(" ", "").replace("\n", "") for c in CATEGORY_LABELS}
    for val in non_empty:
        flat = val.replace(" ", "").replace("\n", "")
        if flat in cat_set:
            return True
        if re.match(r"^[一二三四五六七八九十]+、", flat):
            return True
    return False


def is_skip_row(row):
    non_empty = [clean_flat(c) for c in row if clean_flat(c)]
    if not non_empty:
        return True
    first_few = " ".join(clean_flat(c) for c in row[:4])
    if any(kw in first_few for kw in SKIP_KW):
        return True
    if any("小计" in v for v in non_empty[:3]):
        return True
    return False


def is_header_like(cells):
    non_empty = [c for c in cells if c]
    if not non_empty:
        return True
    flat_cells = [c.replace(" ", "").replace("\n", "") for c in non_empty]
    if all(c in HDR_SET for c in flat_cells):
        return True
    if any(re.match(r"^[一二三四五六七八九十]+、", c) for c in flat_cells):
        return True
    return False


def is_metadata_row(row):
    first = clean_flat(row[0]) if row else ""
    txt = " ".join(clean_flat(c) for c in row)
    if re.match(r"^附件[0-9０１２３]", first):
        return True
    if "填报单位" in txt or "填报时间" in txt:
        return True
    if ("调整表" in txt or "调整情况" in txt or "统计表" in txt) and sum(1 for c in row if clean(c)) <= 3:
        return True
    if "单位：万元" in txt:
        return True
    return False


# ── Semantic header parser ───────────────────────────────────────

def find_header_rows(rows, max_scan=20):
    """
    Find header row(s) and build column mapping.
    Returns (data_start_idx, col_map) where col_map = {src_col_idx: std_col_name}.
    """
    if not rows:
        return -1, {}

    n = min(max_scan, len(rows))

    # Skip metadata rows
    start = 0
    for i in range(n):
        if is_metadata_row(rows[i]):
            start = i + 1
        else:
            break

    best_score = 0
    best_map = {}
    best_data_start = -1

    for i in range(start, n):
        row = rows[i]
        if is_skip_row(row):
            continue

        cleaned = [clean(c) for c in row]

        # Try single-row header
        col_map = _try_map_header(cleaned)
        score = len(col_map)

        # If category row, try multi-row header
        if is_category_row(row) and i + 1 < n:
            merged_map = _try_multi_row_header(rows, i, min(i + 5, n))
            if len(merged_map) > score:
                col_map = merged_map
                score = len(merged_map)
                data_start = i + 1
                for j in range(i + 1, min(i + 6, n)):
                    row_j = [clean(c) for c in rows[j]]
                    if is_header_like(row_j) or is_category_row(rows[j]):
                        data_start = j + 1
                    else:
                        break
                if score > best_score:
                    best_score = score
                    best_map = col_map
                    best_data_start = data_start
                continue

        # Try combining with next row (sub-header pattern)
        if score >= 2 and i + 1 < n:
            sub_row = [clean(c) for c in rows[i + 1]]
            merged = _try_merge_sub_header(cleaned, sub_row)
            if len(merged) > score:
                col_map = merged
                score = len(merged)
                if score > best_score:
                    best_score = score
                    best_map = col_map
                    best_data_start = i + 2
                continue

        if score > best_score:
            best_score = score
            best_map = col_map
            data_start = i + 1
            while data_start < n:
                next_row = [clean(c) for c in rows[data_start]]
                if is_header_like(next_row) or is_category_row(rows[data_start]):
                    data_start += 1
                else:
                    break
            best_data_start = data_start

    return best_data_start, best_map


def _try_map_header(cells):
    col_map = {}
    for i, cell in enumerate(cells):
        resolved = resolve_alias(cell)
        if resolved and resolved in STD_IDX:
            col_map[i] = resolved
    return col_map


def _try_multi_row_header(rows, start, end):
    all_maps = {}
    for i in range(start, end):
        row = [clean(c) for c in rows[i]]
        row_map = _try_map_header(row)
        for idx, name in row_map.items():
            if idx not in all_maps:
                all_maps[idx] = name
    return all_maps


def _try_merge_sub_header(main_row, sub_row):
    """
    Merge main + sub header rows with context tracking.
    E.g., main: [序号, 债券名称, ..., 调整前项目信息, , , 调整后项目信息, , ]
          sub:  [, , ..., 项目名称, 项目单位, 主管部门, 项目名称, 项目单位, 主管部门]
    """
    col_map = {}
    ncols = max(len(main_row), len(sub_row))
    context_prefix = ""

    for i in range(ncols):
        main_val = main_row[i] if i < len(main_row) else ""
        sub_val = sub_row[i] if i < len(sub_row) else ""

        # Update context from main row
        main_flat = main_val.replace("\n", "").replace(" ", "")
        if any(k in main_flat for k in ("调整前", "调出", "变更前", "变动前", "用途调整前")):
            context_prefix = "调整前"
        elif any(k in main_flat for k in ("调整后", "调入", "变更后", "变动后", "用途调整后")):
            context_prefix = "调整后"

        # Try main first
        resolved = resolve_alias(main_val)
        if resolved and resolved in STD_IDX:
            col_map[i] = resolved
            continue

        if sub_val:
            sub_flat = sub_val.replace("\n", "").replace(" ", "")
            # Apply context prefix for ambiguous sub-fields
            ambiguous = {"项目名称", "项目编码", "项目领域", "主管部门", "项目单位",
                         "建设状态", "额度", "金额", "市（州）名称", "县（市、区）名称",
                         "市县名称", "区划编码"}
            if context_prefix and sub_flat in ambiguous:
                contextual = f"{context_prefix}{sub_flat}"
                resolved = resolve_alias(contextual)
                if resolved and resolved in STD_IDX:
                    col_map[i] = resolved
                    continue
            # Also map sub-field contextual names for location fields
            if context_prefix and ("名称" in sub_flat and ("市" in sub_flat or "县" in sub_flat)):
                target = f"市县名称（{context_prefix[-1]}）"  # 调整前/后 -> 前/后
                # More exact mapping
                if context_prefix == "调整前":
                    col_map[i] = "市县名称（调整前）"
                else:
                    col_map[i] = "市县名称（调整后）"
                continue
            if context_prefix and "编码" in sub_flat:
                if context_prefix == "调整前":
                    col_map[i] = "区划编码（调整前）"
                else:
                    col_map[i] = "区划编码（调整后）"
                continue
            # Plain sub-field
            resolved = resolve_alias(sub_val)
            if resolved and resolved in STD_IDX:
                col_map[i] = resolved

    return col_map


# ── Shanxi year-cols special handler ────────────────────────────

def has_year_cols(rows):
    for row in rows[:6]:
        if "需求申报年度" in " ".join(clean_flat(c) for c in row):
            return True
    return False


def remap_shanxi(row):
    cells = [clean_flat(c) for c in row]
    if len(cells) < 32:
        cells.extend([""] * (32 - len(cells)))
    out = cells[:15] + cells[16:23] + cells[24:32]
    if len(out) < 31:
        out.extend([""] * (31 - len(out)))
    return out[:31]


# ── Extract rows with semantic alignment ────────────────────────

def extract_rows_semantic(rows, folder_name, file_name, method):
    """
    Extract data rows using semantic column alignment.
    Returns (output_rows, format_info, quality_info).
    Each output row: [识别方法, 文件夹, 文件名, ...31 STD_COLS...]
    """
    if not rows:
        return [], "empty", {}

    # Shanxi year-cols special case
    is_sx = has_year_cols(rows)
    if is_sx:
        return _extract_shanxi(rows, folder_name, file_name, method)

    # Semantic header parsing
    data_start, col_map = find_header_rows(rows)

    if data_start < 0 or not col_map:
        # Fallback: legacy positional approach
        return _extract_positional_fallback(rows, folder_name, file_name, method)

    ncols_mapped = len(col_map)
    format_tag = f"semantic_{ncols_mapped}cols"

    out = []
    for i in range(data_start, len(rows)):
        row = rows[i]
        if is_skip_row(row):
            continue
        std_row = [""] * 31
        for src_idx, std_name in col_map.items():
            dst_idx = STD_IDX[std_name]
            val = clean_flat(row[src_idx]) if src_idx < len(row) else ""
            std_row[dst_idx] = val

        if is_header_like(std_row):
            continue

        filled = sum(1 for c in std_row if c)
        if filled < 2:
            continue

        out.append([method, folder_name, file_name] + std_row)

    # Merged cell gap-filling
    out = fill_merged_gaps(out)

    quality = assess_quality(out, ncols_mapped)
    return out, format_tag, quality


def _extract_shanxi(rows, folder_name, file_name, method):
    hi = _find_hdr_legacy(rows)
    if hi < 0:
        return [], "shanxi_no_header", {}
    out = []
    for i in range(hi + 1, len(rows)):
        row = rows[i]
        if is_skip_row(row):
            continue
        cells = remap_shanxi(row)
        if is_header_like(cells):
            continue
        if not cells[1] and not cells[10] and not cells[15] and not cells[22]:
            continue
        out.append([method, folder_name, file_name] + cells)
    return out, "shanxi_remap", assess_quality(out, 31)


def _extract_positional_fallback(rows, folder_name, file_name, method):
    hi = _find_hdr_legacy(rows)
    if hi < 0:
        return [], "no_header", {}
    out = []
    for i in range(hi + 1, len(rows)):
        row = rows[i]
        if is_skip_row(row):
            continue
        cells = _norm_row(row)
        if is_header_like(cells):
            continue
        if not cells[1] and not cells[10] and not cells[15] and not cells[22]:
            continue
        out.append([method, folder_name, file_name] + cells)
    return out, "positional_fallback", assess_quality(out, 31)


def _find_hdr_legacy(rows, kw="债券编码", mx=15):
    for i, row in enumerate(rows[:mx]):
        txt = " ".join(clean_flat(c) for c in row)
        if kw in txt:
            return i
    return -1


def _norm_row(row):
    cells = [clean_flat(c) for c in row]
    if len(cells) < 31:
        cells.extend([""] * (31 - len(cells)))
    elif len(cells) > 31:
        cells = cells[:31]
    return cells


# ── Merged cell gap-filling ──────────────────────────────────────

def fill_merged_gaps(rows):
    """
    For rows that continue merged cells (bond info empty, project info present),
    inherit bond info from the row above.
    """
    if len(rows) < 2:
        return rows

    # Bond columns: indices 1-10 in STD_COLS, offset by 3 for method/folder/file prefix
    BOND_COLS = list(range(3 + 1, 3 + 11))
    REGION_COLS = list(range(3 + 11, 3 + 15))

    for i in range(1, len(rows)):
        curr = rows[i]
        prev = rows[i - 1]
        bond_empty = all(not curr[j] for j in BOND_COLS if j < len(curr))
        has_project = any(curr[j] for j in range(3 + 15, min(3 + 28, len(curr))))

        if bond_empty and has_project:
            for j in BOND_COLS:
                if j < len(curr) and j < len(prev) and not curr[j]:
                    curr[j] = prev[j]
            for j in REGION_COLS:
                if j < len(curr) and j < len(prev) and not curr[j]:
                    curr[j] = prev[j]

    return rows


# ── Quality assessment ───────────────────────────────────────────

def assess_quality(rows, mapped_cols):
    if not rows:
        return {"row_count": 0, "quality_grade": "empty"}

    n = len(rows)
    total_empty = 0
    total_cells = 0
    bond_code_count = 0
    project_name_count = 0

    for r in rows:
        data_cells = r[3:] if len(r) > 3 else []
        total_cells += len(data_cells)
        total_empty += sum(1 for c in data_cells if not c)
        if len(data_cells) > 1 and data_cells[1]:
            bond_code_count += 1
        if len(data_cells) > 15 and data_cells[15]:
            project_name_count += 1

    empty_rate = total_empty / total_cells if total_cells > 0 else 1.0

    if mapped_cols >= 20 and empty_rate < 0.5 and (bond_code_count / n) > 0.5:
        grade = "high"
    elif mapped_cols >= 10 and empty_rate < 0.7:
        grade = "medium"
    elif mapped_cols >= 4:
        grade = "low"
    else:
        grade = "very_low"

    return {
        "row_count": n,
        "mapped_cols": mapped_cols,
        "empty_rate": round(empty_rate, 3),
        "has_bond_code": bond_code_count > 0,
        "has_project_name": project_name_count > 0,
        "quality_grade": grade,
    }


# ── File loading ─────────────────────────────────────────────────

def load_excel_rows(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        return [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]

    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


# ── Phase mergers ────────────────────────────────────────────────

def merge_excel_phase(classify, lines):
    rows_all = []
    stats = Counter()
    details = []

    entries = classify.get("excel", [])
    lines.append(f"=== Merge Excel Source ({len(entries)} files) ===")
    for e in entries:
        folder_name = os.path.basename(e["folder"])
        file_name = e["name"]
        try:
            rows = load_excel_rows(e["path"])
            out, fmt, quality = extract_rows_semantic(rows, folder_name, file_name, "Excel表格")
            stats[fmt] += 1
            rows_all.extend(out)
            details.append({
                "folder": folder_name, "file": file_name,
                "format": fmt, "rows": len(out), "quality": quality,
            })
        except Exception as ex:
            stats["error"] += 1
            details.append({
                "folder": folder_name, "file": file_name,
                "format": "error", "rows": 0, "issue": str(ex)[:200],
            })

    lines.append(f"Excel merged rows: {len(rows_all)}")
    lines.append(f"Excel stats: {dict(stats)}")
    return rows_all, details, dict(stats)


def merge_extracted_phase(classify, key_results, base_dir, phase_name, method_label, lines):
    rows_all = []
    stats = Counter()
    details = []

    from openpyxl import load_workbook

    entries = classify.get(key_results, [])
    lines.append(f"\n=== Merge {phase_name} Extracted ({len(entries)} files) ===")

    for e in entries:
        status = e.get("status", "")
        if status not in {"ok", "skipped"}:
            stats["skip_not_ok"] += 1
            continue

        out_name = e.get("out", "")
        if not out_name:
            stats["skip_no_out"] += 1
            continue

        xlsx_path = os.path.join(base_dir, out_name)
        if not os.path.exists(xlsx_path):
            stats["missing_file"] += 1
            details.append({
                "folder": e.get("folder", ""), "file": e.get("file", ""),
                "format": "missing_file", "rows": 0,
            })
            continue

        folder_name = e.get("folder", "")
        file_name = e.get("file", "")

        try:
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)
            file_rows = []
            file_quality = {}
            for ws in wb.worksheets:
                sheet_rows = [list(r) for r in ws.iter_rows(values_only=True)]
                if sheet_rows and sheet_rows[0] and str(sheet_rows[0][0] or "").startswith("# page="):
                    sheet_rows = sheet_rows[1:]
                out, fmt, quality = extract_rows_semantic(
                    sheet_rows, folder_name, file_name, method_label
                )
                file_rows.extend(out)
                if not file_quality or quality.get("row_count", 0) > file_quality.get("row_count", 0):
                    file_quality = quality

            wb.close()
            rows_all.extend(file_rows)
            stats["ok"] += 1
            stats["rows"] += len(file_rows)
            details.append({
                "folder": folder_name, "file": file_name,
                "format": "aligned" if file_rows else "no_data",
                "rows": len(file_rows), "quality": file_quality,
            })
        except Exception as ex:
            stats["error"] += 1
            details.append({
                "folder": folder_name, "file": file_name,
                "format": "error", "rows": 0, "issue": str(ex)[:200],
            })

    lines.append(f"{phase_name} merged rows: {stats.get('rows', 0)}")
    lines.append(f"{phase_name} stats: {dict(stats)}")
    return rows_all, details, dict(stats)


def merge_scanned_phase(classify, lines):
    """Merge scanned PDFs: prefer vision over OCR."""
    rows_all = []
    stats = Counter()
    details = []

    from openpyxl import load_workbook

    ocr_entries = classify.get("scanned_ocr_results", [])
    vision_lookup = {}
    for vr in classify.get("scanned_vision_results", []):
        vk = vr.get("file", "") + "|" + vr.get("folder", "")
        if vr.get("status") == "ok":
            vision_lookup[vk] = vr

    lines.append(f"\n=== Merge Scanned ({len(ocr_entries)} OCR files, {len(vision_lookup)} vision overrides) ===")

    for e in ocr_entries:
        status = e.get("status", "")
        if status not in ("ok", "skipped"):
            stats["skip_not_ok"] += 1
            continue

        out_name = e.get("out", "")
        if not out_name:
            stats["skip_no_out"] += 1
            continue

        folder_name = e.get("folder", "")
        file_name = e.get("file", "")
        key = file_name + "|" + folder_name

        vr = vision_lookup.get(key)
        if vr and vr.get("out"):
            xlsx_path = os.path.join(VISION_DIR, vr["out"])
            source = "vision"
            method_label = "扫描件vision"
        else:
            xlsx_path = os.path.join(SCAN_DIR, out_name)
            source = "ocr"
            method_label = "扫描件OCR"

        if not os.path.exists(xlsx_path):
            stats["missing_file"] += 1
            details.append({
                "folder": folder_name, "file": file_name,
                "format": "missing_file", "source": source, "rows": 0,
            })
            continue

        try:
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)
            file_rows = []
            file_quality = {}
            for ws in wb.worksheets:
                sheet_rows = [list(r) for r in ws.iter_rows(values_only=True)]
                if sheet_rows and sheet_rows[0] and str(sheet_rows[0][0] or "").startswith("# page="):
                    sheet_rows = sheet_rows[1:]
                out, fmt, quality = extract_rows_semantic(
                    sheet_rows, folder_name, file_name, method_label
                )
                file_rows.extend(out)
                if not file_quality or quality.get("row_count", 0) > file_quality.get("row_count", 0):
                    file_quality = quality

            wb.close()
            rows_all.extend(file_rows)
            stats["ok"] += 1
            stats["rows"] += len(file_rows)
            stats[f"source_{source}"] += 1
            details.append({
                "folder": folder_name, "file": file_name,
                "format": "aligned" if file_rows else "no_data",
                "source": source, "rows": len(file_rows),
                "quality": file_quality,
            })
        except Exception as ex:
            stats["error"] += 1
            details.append({
                "folder": folder_name, "file": file_name,
                "format": "error", "source": source, "rows": 0,
                "issue": str(ex)[:200],
            })

    lines.append(f"Scanned merged rows: {stats.get('rows', 0)}")
    lines.append(f"Scanned stats: {dict(stats)}")
    return rows_all, details, dict(stats)


# ── Dedupe & output ──────────────────────────────────────────────

def dedupe_rows(rows):
    seen = set()
    out = []
    for r in rows:
        key = tuple(r)
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def save_output(rows, out_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "资金用途调整表"
    ws.append(["识别方法", "文件夹名称", "表格文件名称"] + STD_COLS)
    for r in rows:
        ws.append(r)
    wb.save(out_path)


# ── CLI ──────────────────────────────────────────────────────────

def parse_args():
    include_scanned = "--include-scanned" in sys.argv
    out = "data/celma-major-events-attachments/merged_用途调整表.xlsx"
    if "--out" in sys.argv:
        i = sys.argv.index("--out")
        if i + 1 < len(sys.argv):
            out = sys.argv[i + 1]
    return include_scanned, out


def main():
    include_scanned, out_path = parse_args()

    with open(CLASSIFY, "r", encoding="utf8") as f:
        classify = json.load(f)

    lines = ["=== Merge From Extracted Tables (v2 - Semantic Alignment) ===",
             f"include_scanned={include_scanned}"]

    excel_rows, excel_details, excel_stats = merge_excel_phase(classify, lines)
    text_rows, text_details, text_stats = merge_extracted_phase(
        classify, "textual_extract_results", TEXT_DIR, "Textual PDF", "标准件提取", lines
    )

    scan_rows = []
    scan_details = []
    scan_stats = {}
    if include_scanned:
        scan_rows, scan_details, scan_stats = merge_scanned_phase(classify, lines)

    all_rows = excel_rows + text_rows + scan_rows
    before = len(all_rows)
    all_rows = dedupe_rows(all_rows)
    deduped = before - len(all_rows)

    save_output(all_rows, out_path)

    # Quality summary
    quality_summary = {"high": 0, "medium": 0, "low": 0, "very_low": 0, "empty": 0}
    for phase_details in [excel_details, text_details, scan_details]:
        for d in phase_details:
            q = d.get("quality", {})
            grade = q.get("quality_grade", "empty")
            quality_summary[grade] = quality_summary.get(grade, 0) + 1

    classify["merge_from_extracted_summary"] = {
        "version": "v2_semantic",
        "include_scanned": include_scanned,
        "excel_rows": len(excel_rows),
        "textual_rows": len(text_rows),
        "scanned_rows": len(scan_rows),
        "total_before_dedupe": before,
        "deduped_rows": deduped,
        "total_rows": len(all_rows),
        "quality_summary": quality_summary,
        "output": out_path,
    }
    classify["merge_from_extracted_details"] = {
        "excel": {"stats": excel_stats, "files": excel_details},
        "textual": {"stats": text_stats, "files": text_details},
        "scanned": {"stats": scan_stats, "files": scan_details},
    }
    with open(CLASSIFY, "w", encoding="utf8") as f:
        json.dump(classify, f, ensure_ascii=False, indent=2)

    lines.append("\n=== Quality Summary ===")
    for grade, count in sorted(quality_summary.items()):
        lines.append(f"  {grade}: {count} files")

    lines.append("\n=== Summary ===")
    lines.append(f"Output: {out_path}")
    lines.append(f"Excel rows: {len(excel_rows)}")
    lines.append(f"Textual rows: {len(text_rows)}")
    lines.append(f"Scanned rows: {len(scan_rows)}")
    lines.append(f"Rows before dedupe: {before}")
    lines.append(f"Deduped: {deduped}")
    lines.append(f"Rows final: {len(all_rows)}")

    # Report files with 0 rows for debugging
    zero_files = []
    for phase_name, phase_details in [("excel", excel_details), ("textual", text_details), ("scanned", scan_details)]:
        for d in phase_details:
            if d.get("rows", 0) == 0 and d.get("format") not in ("missing_file", "error"):
                zero_files.append(f"  [{phase_name}] {d.get('file', '')[:50]} -> {d.get('format', '')}")

    if zero_files:
        lines.append(f"\n=== Zero-row files ({len(zero_files)}) ===")
        for zf in zero_files:
            lines.append(zf)

    with open(LOG_FILE, "w", encoding="utf8") as f:
        f.write("\n".join(lines))

    print(f"Done. rows={len(all_rows)} deduped={deduped}")
    print(f"Quality: {quality_summary}")
    if zero_files:
        print(f"Zero-row files: {len(zero_files)}")
    print(f"See {LOG_FILE}")


if __name__ == "__main__":
    main()
